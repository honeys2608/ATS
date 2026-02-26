# app/routes/candidates.py



from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query

from sqlalchemy.orm import Session

from typing import List, Optional

from datetime import datetime

import os

import uuid

import json

import re
import asyncio

from fastapi.responses import FileResponse, StreamingResponse

from sqlalchemy import func, String, cast, or_, and_

from sqlalchemy.exc import IntegrityError

from io import BytesIO

import pandas as pd







from app.db import get_db
from app.db import SessionLocal

from app import models, schemas

from app.ai_core import (

    generate_candidate_embedding,

    calculate_fit_score,

    generate_fit_explanation,

    add_to_faiss_index,

)

from app.resume_parser import parse_resume as parse_resume_file

from app.auth import get_current_user

from app.permissions import require_permission

from app.utils.role_check import allow_user

from app.utils.candidate_bulk_upload import (

    ALLOWED_EXTENSION,

    ALLOWED_MIME_TYPES,

    MAX_FILE_SIZE_BYTES,

    MAX_ROWS,

    TEMPLATE_HEADERS,

    validate_headers,

    validate_row,

    parse_email,

)



# Import system settings

from app.models import SystemSettings
from app.task_queue import task_queue, get_task_status, TaskStatus



router = APIRouter(prefix="/v1/candidates", tags=["Admin Candidates"])

UPLOAD_DIR = "uploads/resumes"

os.makedirs(UPLOAD_DIR, exist_ok=True)

BULK_UPLOAD_DIR = "uploads/bulk_uploads"

ERRORS_DIR = os.path.join(BULK_UPLOAD_DIR, "errors")

os.makedirs(ERRORS_DIR, exist_ok=True)

LOCK_RELEASE_STATUSES = {
    "withdrawn",
    "rejected",
    "rejected_by_recruiter",
    "am_rejected",
    "client_rejected",
    "hired",
    "joined",
    "offer_declined",
}


def _normalize_text(value):
    return str(value or "").strip().lower()


def _is_recruiter(current_user):
    return _normalize_text(current_user.get("role")) == "recruiter"


def _resolve_job_by_ref(db: Session, job_ref: Optional[str]):
    if not job_ref:
        return None
    raw = str(job_ref).strip()
    if not raw:
        return None
    return db.query(models.Job).filter(
        (models.Job.id == raw) | (models.Job.job_id == raw)
    ).first()


def _get_user_name(db: Session, user_id: Optional[str]) -> str:
    if not user_id:
        return ""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        return str(user_id)
    return user.full_name or user.username or user.email or user.id


def _ensure_recruiter_assigned_to_job(db: Session, current_user, job: models.Job):
    if not _is_recruiter(current_user):
        return
    assigned = db.query(models.job_recruiters).filter(
        models.job_recruiters.c.job_id == job.id,
        models.job_recruiters.c.recruiter_id == current_user.get("id"),
    ).first()
    if not assigned:
        raise HTTPException(
            status_code=403,
            detail=f"You are not assigned to job {job.job_id or job.id}",
        )


def _enforce_recruiter_candidate_job_lock(
    db: Session,
    *,
    current_user,
    candidate_id: str,
    job_ref: Optional[str],
    action: str,
):
    if not _is_recruiter(current_user):
        return None
    job = _resolve_job_by_ref(db, job_ref)
    if not job:
        raise HTTPException(400, "job_id is required for recruiter candidate edits")
    _ensure_recruiter_assigned_to_job(db, current_user, job)

    submission = db.query(models.CandidateSubmission).filter(
        models.CandidateSubmission.candidate_id == candidate_id,
        models.CandidateSubmission.job_id == job.id,
    ).first()
    recruiter_id = current_user.get("id")

    if submission:
        if (
            submission.recruiter_id
            and str(submission.recruiter_id) != str(recruiter_id)
            and submission.is_locked
            and _normalize_text(submission.status) not in LOCK_RELEASE_STATUSES
        ):
            owner_name = _get_user_name(db, submission.recruiter_id)
            raise HTTPException(
                status_code=409,
                detail={
                    "message": f"Candidate is in progress by recruiter {owner_name}",
                    "lock_status_label": f"In Progress by Recruiter {owner_name}",
                    "assignment_label": f"Assigned to Recruiter {owner_name} for Job ID {job.job_id or job.id}",
                    "locked_by_recruiter_id": submission.recruiter_id,
                    "job_id": job.job_id or job.id,
                    "candidate_id": candidate_id,
                },
            )
        submission.recruiter_id = recruiter_id
        submission.is_locked = True
        submission.stage = submission.stage or "recruiter_review"
        submission.updated_at = datetime.utcnow()
        return job

    submission = models.CandidateSubmission(
        id=models.generate_uuid(),
        candidate_id=candidate_id,
        job_id=job.id,
        recruiter_id=recruiter_id,
        status="submitted",
        stage="recruiter_review",
        is_locked=True,
        source=action,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(submission)
    return job







def clean_null(value):

    if value in ["null", "None", "", [], {}]:

        return None

    return value





# ============================================================

# Candidate Bulk Upload - Template Download

# ============================================================

@router.get("/bulk-template")

@require_permission("candidates", "view")

async def download_candidate_bulk_template(

    current_user=Depends(get_current_user),

):

    allow_user(current_user)

    try:

        import openpyxl

    except ImportError:

        raise HTTPException(500, "openpyxl not installed")



    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Candidates"

    ws.append(TEMPLATE_HEADERS)

    ws.append(

        [

            "Alex Morgan",

            "alex.morgan@gmail.com",

            "9876543210",

            "9123456789",

            "1995-04-12",

            "Female",

            "Single",

            "Acme Corp",

            "Senior Developer",

            "6",

            "4",

            "12.5",

            "15",

            "30",

            "Bengaluru",

            "Remote",

            "Python, React",

            "Python",

            "React",

            "B.Tech",

            "State University",

            "2017",

            "AWS Certified Developer",

            "https://example.com/resume.pdf",

            "https://linkedin.com/in/alexmorgan",

            "https://github.com/alexmorgan",

            "https://alexmorgan.dev",

            "123 Main Street",

            "Bengaluru",

            "Karnataka",

            "India",

            "560001",

            "Yes",

            "Full-time",

            "2026-02-15",

            "2026-02-01",

            "Strong backend experience",

            "Referral",

        ]

    )



    output = BytesIO()

    wb.save(output)

    output.seek(0)



    headers = {"Content-Disposition": "attachment; filename=Candidate_Bulk_Upload_Template.xlsx"}

    return StreamingResponse(

        output,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers=headers,

    )





# ============================================================

# Candidate Bulk Upload - History

# ============================================================

@router.get("/bulk-upload/history")

@require_permission("candidates", "view")

async def get_candidate_bulk_upload_history(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)



    logs = (

        db.query(models.CandidateBulkUploadLog)

        .order_by(models.CandidateBulkUploadLog.created_at.desc())

        .limit(50)

        .all()

    )



    response = []

    for log in logs:

        response.append(

            schemas.CandidateBulkUploadLogResponse(

                id=log.id,

                filename=log.filename,

                total_rows=log.total_rows,

                success_count=log.success_count,

                failed_count=log.failed_count,

                error_csv_url=f"/v1/candidates/bulk-upload/errors/{log.id}"

                if log.error_csv_path

                else None,

                created_at=log.created_at,

            )

        )

    return response





# ============================================================

# Candidate Bulk Upload - Error CSV Download

# ============================================================

@router.get("/bulk-upload/errors/{log_id}")

@require_permission("candidates", "view")

async def download_candidate_bulk_upload_errors(

    log_id: str,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)

    log = db.query(models.CandidateBulkUploadLog).filter(models.CandidateBulkUploadLog.id == log_id).first()

    if not log or not log.error_csv_path:

        raise HTTPException(404, "Error CSV not found")

    if not os.path.exists(log.error_csv_path):

        raise HTTPException(404, "Error CSV not found")

    return FileResponse(

        log.error_csv_path,

        media_type="text/csv",

        filename=f"candidate_bulk_upload_errors_{log_id}.csv",

    )





# ============================================================

# Candidate Bulk Upload - Upload Endpoint

# ============================================================

@router.post("/bulk-upload")

@require_permission("candidates", "create")

async def bulk_upload_candidates(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)



    if not file:

        raise HTTPException(400, "File is required")



    filename = file.filename or "upload.xlsx"

    if not filename.lower().endswith(ALLOWED_EXTENSION):

        raise HTTPException(400, "Only .xlsx files are allowed")



    if file.content_type and file.content_type not in ALLOWED_MIME_TYPES:

        raise HTTPException(400, "Unsupported file type")



    content = await file.read()

    if len(content) > MAX_FILE_SIZE_BYTES:

        raise HTTPException(400, "File size exceeds 10MB limit")



    try:

        df = pd.read_excel(BytesIO(content), dtype=str)

    except Exception:

        raise HTTPException(400, "Invalid Excel file")



    if df.empty:

        raise HTTPException(400, "File is empty")



    if len(df.index) > MAX_ROWS:

        raise HTTPException(400, "Row limit exceeded (max 5000)")



    headers = list(df.columns)

    valid, missing, extra, ordered_ok = validate_headers(headers)

    if not valid or not ordered_ok:

        raise HTTPException(

            400,

            {

                "message": "Invalid template headers",

                "missing": missing,

                "extra": extra,

                "ordered": ordered_ok,

                "expected": TEMPLATE_HEADERS,

            },

        )



    raw_rows = df.fillna("").to_dict(orient="records")



    parsed_emails = []

    for row in raw_rows:

        email = parse_email(row.get("Email*"))

        if email:

            parsed_emails.append(email)

    email_set = set(parsed_emails)

    existing_emails = set()

    if email_set:

        existing = (

            db.query(models.Candidate.email)

            .filter(models.Candidate.email.in_(email_set))

            .all()

        )

        existing_emails = {e[0].lower() for e in existing if e[0]}



    errors: List[dict] = []

    success_count = 0

    failed_count = 0

    seen_emails: set[str] = set()

    for idx, row in enumerate(raw_rows):

        row_number = idx + 2  # account for header row

        result = validate_row(row, row_number, existing_emails, seen_emails)

        if result.errors:

            errors.extend(result.errors)

            failed_count += 1

            continue



        payload = result.data



        notice_days = payload.get("notice_period_days")

        if notice_days is not None:

            payload["notice_period"] = str(notice_days)



        candidate = models.Candidate(

            full_name=payload.get("full_name"),

            email=payload.get("email"),

            phone=payload.get("phone"),

            alternate_phone=payload.get("alternate_phone"),

            dob=payload.get("dob"),

            gender=payload.get("gender"),

            marital_status=payload.get("marital_status"),

            current_employer=payload.get("current_employer"),

            current_job_title=payload.get("current_job_title"),

            experience_years=payload.get("experience_years"),

            relevant_experience_years=payload.get("relevant_experience_years"),

            current_ctc=payload.get("current_ctc"),

            expected_ctc=payload.get("expected_ctc"),

            notice_period=payload.get("notice_period"),

            notice_period_days=payload.get("notice_period_days"),

            current_location=payload.get("current_location"),

            preferred_location=payload.get("preferred_location"),

            skills=payload.get("skills"),

            primary_skill=payload.get("primary_skill"),

            secondary_skill=payload.get("secondary_skill"),

            qualification=payload.get("qualification"),

            university=payload.get("university"),

            graduation_year=payload.get("graduation_year"),

            certifications_text=payload.get("certifications_text"),

            resume_url=payload.get("resume_url"),

            linkedin_url=payload.get("linkedin_url"),

            github_url=payload.get("github_url"),

            portfolio_url=payload.get("portfolio_url"),

            current_address=payload.get("current_address"),

            city=payload.get("city"),

            state=payload.get("state"),

            country=payload.get("country"),

            pincode=payload.get("pincode"),

            willing_to_relocate=payload.get("willing_to_relocate"),

            preferred_employment_type=payload.get("preferred_employment_type"),

            availability_to_join=payload.get("availability_to_join"),

            last_working_day=payload.get("last_working_day"),

            internal_notes=payload.get("internal_notes"),

            source=payload.get("source") or "Excel Upload",

            status=models.CandidateStatus.sourced,

            public_id=models.generate_candidate_public_id_from_org(db),

        )

        try:

            db.add(candidate)

            db.flush()

            success_count += 1

        except IntegrityError:

            db.rollback()

            errors.append(

                {

                    "row": row_number,

                    "field": "email",

                    "message": "Email violates database constraints",

                }

            )

            failed_count += 1



    db.commit()



    error_csv_path = None

    if errors:

        error_csv_path = os.path.join(ERRORS_DIR, f"{uuid.uuid4()}.csv")

        pd.DataFrame(errors).to_csv(error_csv_path, index=False)



    log = models.CandidateBulkUploadLog(

        uploaded_by=current_user.get("id"),

        filename=filename,

        total_rows=len(raw_rows),

        success_count=success_count,

        failed_count=failed_count,

        error_csv_path=error_csv_path,

    )

    db.add(log)

    db.commit()



    return {

        "total_rows": len(raw_rows),

        "success_count": success_count,

        "failed_count": failed_count,

        "errors": errors,

    }



# ============================================================

# ⭐ HELPER — Generate Candidate ID {ORG}-C-0001

# ============================================================

def generate_candidate_public_id_from_org(db: Session, org_code: str) -> str:



    if not org_code or not isinstance(org_code, str):

        org_code = "ORG"



    org_code = org_code.strip().upper()

    if len(org_code) != 3 or not org_code.isalpha():

        org_code = "ORG"



    prefix = f"{org_code}-C-"



    existing_ids = (

        db.query(models.Candidate.public_id)

        .filter(models.Candidate.public_id.like(f"{prefix}%"))

        .all()

    )



    max_num = 0

    for (pid,) in existing_ids:

        try:

            num = int(pid.replace(prefix, ""))

            max_num = max(max_num, num)

        except:

            continue



    next_num = max_num + 1



    return f"{prefix}{next_num:04d}"









# ============================================================

# ⭐ HELPER — Normalize candidate for response

# ============================================================

def normalize_candidate(candidate: models.Candidate):

    """

    Converts AI / parsed fields into frontend-safe primitives

    """



    if hasattr(candidate, "education") and isinstance(candidate.education, dict):

        candidate.education = candidate.education.get("raw")



    if hasattr(candidate, "experience") and isinstance(candidate.experience, dict):

        candidate.experience = candidate.experience.get("raw")



    if hasattr(candidate, "skills") and isinstance(candidate.skills, dict):

        candidate.skills = candidate.skills.get("raw")



    return candidate


def resolve_candidate_job_account_manager(
    db: Session,
    candidate: models.Candidate,
):
    mapped_job = None

    if candidate.applied_job_id:
        mapped_job = (
            db.query(models.Job)
            .filter(models.Job.id == candidate.applied_job_id)
            .first()
        )

    if not mapped_job:
        latest_application = (
            db.query(models.JobApplication.job_id)
            .filter(models.JobApplication.candidate_id == candidate.id)
            .order_by(
                models.JobApplication.applied_at.desc(),
                models.JobApplication.id.desc(),
            )
            .first()
        )
        latest_job_id = latest_application[0] if latest_application else None
        if latest_job_id:
            mapped_job = (
                db.query(models.Job)
                .filter(models.Job.id == latest_job_id)
                .first()
            )
            if mapped_job and not candidate.applied_job_id:
                candidate.applied_job_id = mapped_job.id

    if not mapped_job:
        return None

    am_user = mapped_job.account_manager
    return {
        "job_id": mapped_job.id,
        "job_title": mapped_job.title,
        "account_manager_id": mapped_job.account_manager_id,
        "account_manager_name": am_user.full_name if am_user else None,
        "account_manager_email": am_user.email if am_user else None,
    }


@router.get("")
async def list_candidates(
    status: Optional[str] = None,
    source: Optional[str] = None,
    q: Optional[str] = None,
    search: Optional[str] = None,
    applied_job: Optional[str] = None,
    page: Optional[int] = Query(None, ge=1),
    limit: int = Query(9, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    allow_user(current_user)



    query = (

        db.query(

            models.Candidate,

            func.max(models.Job.id).label("job_id"),

            func.max(models.Job.title).label("job_title"),

            func.max(models.Job.account_manager_id).label("account_manager_id"),

            func.max(models.User.full_name).label("account_manager_name"),

            func.max(models.User.email).label("account_manager_email"),

        )

        # ✅ LEFT JOIN so bulk candidates without jobs are included

        .outerjoin(

            models.JobApplication,

            models.JobApplication.candidate_id == models.Candidate.id

        )

        .outerjoin(

            models.Job,

            models.Job.id == models.JobApplication.job_id

        )

        .outerjoin(

            models.User,

            models.User.id == models.Job.account_manager_id

        )

        .filter(models.Candidate.merged_into_id.is_(None))

    )



    if status:

        query = query.filter(models.Candidate.status == status)



    if source:

        query = query.filter(

            func.lower(models.Candidate.source) == source.lower()

        )



    # Allow "search" alias for q (used by Resdex Send NVite)

    if not q and search:

        q = search



    if q:

        like = f"%{q}%"

        query = query.filter(

            models.Candidate.full_name.ilike(like) |

            models.Candidate.email.ilike(like)

        )



    # Applied job filter (by job title or job id)

    if applied_job:

        job_like = f"%{applied_job}%"

        query = query.filter(

            (models.Job.title.ilike(job_like)) |

            (models.Job.id == applied_job)

        )



    grouped_query = query.group_by(models.Candidate.id)
    if page is None:
        results = grouped_query.order_by(models.Candidate.created_at.desc()).all()
        total_records = len(results)
    else:
        total_records = grouped_query.count()
        results = (
            grouped_query
            .order_by(models.Candidate.created_at.desc())
            .offset((page - 1) * limit)
            .limit(limit)
            .all()
        )



    current_user_id = str(current_user.get("id") or "").strip()
    candidate_rows = list(results)
    candidate_ids = [str(row[0].id) for row in candidate_rows if row and row[0] and row[0].id]

    ownership_by_candidate = {}
    if candidate_ids:
        app_rows = (
            db.query(models.JobApplication)
            .filter(
                models.JobApplication.candidate_id.in_(candidate_ids),
                models.JobApplication.recruiter_id.isnot(None),
            )
            .all()
        )
        for app in app_rows:
            status_key = _normalize_text(app.status)
            if status_key in LOCK_RELEASE_STATUSES:
                continue
            owner_id = str(app.recruiter_id or "").strip()
            if not owner_id:
                continue
            ts = (
                app.decision_at
                or app.sent_to_client_at
                or app.sent_to_am_at
                or app.shortlisted_at
                or app.applied_at
                or getattr(app, "created_at", None)
                or datetime.min
            )
            existing = ownership_by_candidate.get(app.candidate_id)
            if not existing or ts >= existing.get("timestamp", datetime.min):
                ownership_by_candidate[app.candidate_id] = {
                    "recruiter_id": owner_id,
                    "timestamp": ts,
                    "status": app.status,
                }

        submission_rows = (
            db.query(models.CandidateSubmission)
            .filter(
                models.CandidateSubmission.candidate_id.in_(candidate_ids),
                models.CandidateSubmission.recruiter_id.isnot(None),
            )
            .all()
        )
        for sub in submission_rows:
            status_key = _normalize_text(sub.status)
            if status_key in LOCK_RELEASE_STATUSES:
                continue
            if sub.is_locked is False:
                continue
            owner_id = str(sub.recruiter_id or "").strip()
            if not owner_id:
                continue
            ts = sub.updated_at or sub.created_at or datetime.min
            existing = ownership_by_candidate.get(sub.candidate_id)
            if not existing or ts >= existing.get("timestamp", datetime.min):
                ownership_by_candidate[sub.candidate_id] = {
                    "recruiter_id": owner_id,
                    "timestamp": ts,
                    "status": sub.status or sub.stage,
                }

    recruiter_name_by_id = {}
    recruiter_role_by_id = {}
    owner_ids = list(
        {
            str(record.get("recruiter_id") or "").strip()
            for record in ownership_by_candidate.values()
            if record.get("recruiter_id")
        }
    )
    if owner_ids:
        recruiter_rows = (
            db.query(
                models.User.id,
                models.User.full_name,
                models.User.email,
                models.User.username,
                models.User.role,
            )
            .filter(models.User.id.in_(owner_ids))
            .all()
        )
        recruiter_name_by_id = {
            str(row.id): (
                row.full_name
                or row.email
                or row.username
                or str(row.id)
            )
            for row in recruiter_rows
        }
        recruiter_role_by_id = {
            str(row.id): (
                str(row.role.value if hasattr(row.role, "value") else row.role or "")
                .strip()
                .replace("_", " ")
                .title()
            )
            for row in recruiter_rows
        }

    response = []

    for (
        cand,
        job_id,
        job_title,
        account_manager_id,
        account_manager_name,
        account_manager_email,
    ) in candidate_rows:

        c = normalize_candidate(cand)
        resolved_mapping = resolve_candidate_job_account_manager(db, c)

        existing_job_title = getattr(c, "job_title", None)

        if resolved_mapping:
            if resolved_mapping["job_id"] and not c.applied_job_id:
                c.applied_job_id = resolved_mapping["job_id"]
            c.job_title = resolved_mapping["job_title"] or existing_job_title
            c.account_manager_id = resolved_mapping["account_manager_id"]
            c.account_manager_name = resolved_mapping["account_manager_name"]
            c.account_manager_email = resolved_mapping["account_manager_email"]
        else:
            if job_id and not c.applied_job_id:
                c.applied_job_id = job_id
            c.job_title = job_title or existing_job_title
            c.account_manager_id = account_manager_id
            c.account_manager_name = account_manager_name
            c.account_manager_email = account_manager_email

        ownership = ownership_by_candidate.get(c.id) or {}
        assigned_recruiter_id = str(ownership.get("recruiter_id") or "").strip()
        assigned_recruiter_name = recruiter_name_by_id.get(assigned_recruiter_id, None)
        assigned_recruiter_role = recruiter_role_by_id.get(assigned_recruiter_id, None)
        assigned_workflow_status = str(ownership.get("status") or "").strip() or None
        c.assigned_recruiter_id = assigned_recruiter_id or None
        c.assigned_recruiter_name = assigned_recruiter_name
        c.assigned_recruiter_role = assigned_recruiter_role
        c.assigned_workflow_status = assigned_workflow_status
        c.is_owned_by_current_user = (
            True if not assigned_recruiter_id else assigned_recruiter_id == current_user_id
        )
        c.is_in_use_by_another_recruiter = bool(
            assigned_recruiter_id and assigned_recruiter_id != current_user_id
        )

        response.append(c)



    if page is None:
        return response

    total_pages = max(1, (total_records + limit - 1) // limit) if total_records else 1
    return {
        "data": response,
        "results": response,
        "currentPage": page,
        "totalPages": total_pages,
        "totalRecords": total_records,
        "count": total_records,
        "limit": limit,
    }

@router.get("/search")

@require_permission("candidates", "view")

async def search_candidates(

    keyword: Optional[str] = None,

    logic: str = "OR",

    min_exp: Optional[float] = None,

    max_exp: Optional[float] = None,

    location: Optional[str] = None,

    limit: int = 20,

    offset: int = 0,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """

    Simple search for candidates with basic filters

    """

    allow_user(current_user)

    

    try:

        # Start with base query - only non-merged candidates

        query = db.query(models.Candidate)

        

        # Filter out merged candidates

        if hasattr(models.Candidate, 'merged_into_id'):

            query = query.filter(models.Candidate.merged_into_id.is_(None))

        

        # Keyword filtering

        if keyword and keyword.strip():

            search_term = f"%{keyword.strip()}%"

            

            # Build search conditions

            search_conditions = [

                models.Candidate.full_name.ilike(search_term),

                models.Candidate.current_employer.ilike(search_term),

                cast(models.Candidate.skills, String).ilike(search_term),

            ]

            

            # Add optional fields if they exist

            if hasattr(models.Candidate, 'experience'):

                search_conditions.append(models.Candidate.experience.ilike(search_term))

            if hasattr(models.Candidate, 'parsed_resume'):

                search_conditions.append(cast(models.Candidate.parsed_resume, String).ilike(search_term))

            

            # Apply logic

            if logic.upper() == "AND":

                query = query.filter(and_(*search_conditions))

            else:  # OR (default)

                query = query.filter(or_(*search_conditions))

        

        # Experience filter

        if min_exp is not None and hasattr(models.Candidate, 'experience_years'):

            query = query.filter(models.Candidate.experience_years >= min_exp)

        if max_exp is not None and hasattr(models.Candidate, 'experience_years'):

            query = query.filter(models.Candidate.experience_years <= max_exp)

        

        # Location filter

        if location and location.strip():

            search_term = f"%{location.strip()}%"

            query = query.filter(models.Candidate.current_location.ilike(search_term))

        

        # Order by creation date descending

        query = query.order_by(models.Candidate.created_at.desc())

        

        # Get total count before pagination

        total_count = query.count()

        

        # Apply pagination

        candidates = query.offset(offset).limit(limit).all()

        

        # Build results

        results = []

        for c in candidates:

            result = {

                "id": c.id,

                "name": c.full_name or "N/A",

                "email": c.email or "N/A",

                "phone": c.phone or "N/A",

                "skills": c.skills if isinstance(c.skills, list) else [],

                "experience": c.experience_years or 0,

                "location": c.current_location or "N/A",

                "city": c.city or "N/A",

                "employer": c.current_employer or "N/A",

                "designation": c.experience or "N/A",

                "salary": c.expected_salary or 0,

                "status": str(c.status) if c.status else "N/A",

                "resume_url": c.resume_url,

            }

            results.append(result)

        

        return {

            "total": total_count,

            "count": len(candidates),

            "limit": limit,

            "offset": offset,

            "results": results,

        }

    

    except Exception as e:

        print(f"Search error: {str(e)}")

        # Return empty results on error instead of crashing

        return {

            "total": 0,

            "count": 0,

            "limit": limit,

            "offset": offset,

            "results": [],

            "error": str(e),

        }



@router.post("/manual")

@require_permission("candidates", "update")

async def manual_add_candidate(

    job_id: str = Form(...),

    full_name: Optional[str] = Form(None),

    email: Optional[str] = Form(None),

    phone: Optional[str] = Form(None),

    resume: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)

    public_id_pattern = re.compile(r"^[A-Z]{3}-C-\d{4}$")



    # 1️⃣ Validate Job

    job = db.query(models.Job).filter(

        (models.Job.job_id == job_id) | (models.Job.id == job_id)

    ).first()



    if not job:

        raise HTTPException(400, "Invalid job_id")
    _ensure_recruiter_assigned_to_job(db, current_user, job)



    # 2️⃣ Save Resume

    filename = f"{uuid.uuid4()}_{resume.filename}"

    file_path = os.path.join(UPLOAD_DIR, filename)



    with open(file_path, "wb") as f:

        f.write(await resume.read())



    # 3️⃣ Parse Resume

    parse_result = parse_resume_file(file_path) or {}

    parsed = parse_result.get("data", {}) if parse_result.get("success") else {}



    parsed_full_name = parsed.get("full_name")

    parsed_email = parsed.get("email")

    parsed_phone = parsed.get("phone")



    # 4️⃣ SAFE FALLBACKS (ATS-CORRECT)

    full_name = full_name or parsed_full_name or "Unknown"

    email = (email or parsed_email or "").strip() or None

    phone = (phone or parsed_phone or "").strip() or None



    # Enforce identity fields to satisfy DB constraint

    if not email and not phone:

        raise HTTPException(

            status_code=400,

            detail="Email or phone number is required to create candidate",

        )



    # 5️⃣ Skills

    skills = parsed.get("skills")

    if not isinstance(skills, list):

        skills = []



    # 6️⃣ Education (ALWAYS LIST)

    education = parsed.get("education") or []

    if isinstance(education, str):

        education = [{"raw": education}]



    # 7️⃣ Create or Update Candidate (avoid duplicate email)

    candidate = None

    if email:

        candidate = db.query(models.Candidate).filter(models.Candidate.email == email).first()

    if not candidate and phone:

        candidate = db.query(models.Candidate).filter(models.Candidate.phone == phone).first()



    if candidate:

        candidate.full_name = full_name or candidate.full_name

        candidate.phone = phone or candidate.phone

        candidate.skills = skills or candidate.skills

        candidate.education = education or candidate.education

        candidate.experience_years = parsed.get("experience_years") or candidate.experience_years or 0

        candidate.resume_url = file_path or candidate.resume_url

        if not candidate.source:

            candidate.source = "recruiter"

        if candidate.status != "verified":

            candidate.status = "verified"

        if not candidate.public_id or not public_id_pattern.match(candidate.public_id):

            candidate.public_id = models.generate_candidate_public_id_from_org(db)

        db.commit()

        db.refresh(candidate)

    else:

        candidate = models.Candidate(

            full_name=full_name,

            email=email,

            phone=phone,

            skills=skills,

            education=education,

            experience_years=parsed.get("experience_years") or 0,

            resume_url=file_path,

            status="verified",

            source="recruiter",

            created_at=datetime.utcnow(),

            public_id=models.generate_candidate_public_id_from_org(db),

        )



        db.add(candidate)

        db.commit()

        db.refresh(candidate)

    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current_user,
        candidate_id=candidate.id,
        job_ref=job.id,
        action="manual_add_candidate",
    )



    # 8️⃣ Create Submission (avoid duplicates)

    submission = db.query(models.CandidateSubmission).filter(

        models.CandidateSubmission.candidate_id == candidate.id,

        models.CandidateSubmission.job_id == job.id

    ).first()



    if not submission:

        submission = models.CandidateSubmission(

            candidate_id=candidate.id,

            job_id=job.id,

            recruiter_id=current_user["id"],

            match_score=0,

            status="submitted",
            stage="recruiter_review",
            is_locked=True,

            source="recruiter",

            created_at=datetime.utcnow(),

            submitted_at=datetime.utcnow(),

        )

        db.add(submission)

        db.commit()
    else:
        submission.recruiter_id = current_user["id"]
        submission.is_locked = True
        submission.stage = submission.stage or "recruiter_review"
        submission.updated_at = datetime.utcnow()
        db.commit()



    # ? Create JobApplication so it appears in View Submissions

    existing_app = db.query(models.JobApplication).filter(

        models.JobApplication.job_id == job.id,

        models.JobApplication.candidate_id == candidate.id

    ).first()



    if not existing_app:

        app = models.JobApplication(

            job_id=job.id,

            candidate_id=candidate.id,

            full_name=candidate.full_name,

            email=candidate.email,

            phone=candidate.phone,

            status="sent_to_am",

            applied_at=datetime.utcnow(),

            sent_to_am_at=datetime.utcnow(),

            recruiter_id=current_user["id"],

        )

        db.add(app)

        db.commit()



# 9️⃣ Response

    return {

        "message": "Candidate added successfully",

        "warnings": [] if email else ["Email not detected in resume"],

        "candidate_id": candidate.id,

        "job_id": job.job_id,

        "parsed": {

            "full_name": full_name,

            "email": email,

            "phone": phone,

            "skills": skills,

            "experience_years": candidate.experience_years,

            "education": education,

            "location": parsed.get("location") or "",

            "resume_text": parsed.get("resume_text") or "",

        },

    }





@router.get("/pool", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def candidate_pool(

    q: Optional[str] = None,

    classification: Optional[str] = None,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)



    query = db.query(models.Candidate).filter(

        models.Candidate.merged_into_id.is_(None)

    )





    if q:

        like = f"%{q}%"

        query = query.filter(

            models.Candidate.full_name.ilike(like) |

            models.Candidate.email.ilike(like)

        )



    if classification:

        query = query.filter(models.Candidate.classification == classification)



    candidates = query.order_by(models.Candidate.created_at.desc()).all()

    return [normalize_candidate(c) for c in candidates]





# ============================================================

# Candidate Bulk Verify (Move to Pool)

# ============================================================

@router.post("/verify-bulk")

@require_permission("candidates", "update")

async def verify_candidates_bulk(

    payload: schemas.CandidateBulkVerifyRequest,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)



    ids = [cid for cid in (payload.ids or []) if cid]

    if not ids:

        raise HTTPException(400, "No candidate IDs provided")



    candidates = (

        db.query(models.Candidate)

        .filter(models.Candidate.id.in_(ids))

        .all()

    )



    if not candidates:

        raise HTTPException(404, "No candidates found")



    updated = 0

    for candidate in candidates:

        candidate.status = "verified"

        updated += 1



    db.commit()



    return {"updated": updated, "status": "verified"}





# ============================================================

# PIPELINE STAGE ENDPOINTS - Filter candidates by status

# ============================================================

@router.get("/stage/sourced", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def get_sourced_candidates(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """Get all candidates in Sourced stage (status='sourced')."""

    allow_user(current_user)

    

    query = db.query(models.Candidate).filter(

        models.Candidate.status == "sourced",

        models.Candidate.merged_into_id.is_(None)

    ).order_by(models.Candidate.created_at.desc())

    

    candidates = query.all()

    return [normalize_candidate(c) for c in candidates]





@router.get("/stage/screening", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def get_screening_candidates(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """Get all candidates in Screening stage (status='screening')."""

    allow_user(current_user)

    

    query = db.query(models.Candidate).filter(

        models.Candidate.status == "screening",

        models.Candidate.merged_into_id.is_(None)

    ).order_by(models.Candidate.created_at.desc())

    

    candidates = query.all()

    return [normalize_candidate(c) for c in candidates]





@router.get("/stage/submitted", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def get_submitted_candidates(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """Get all candidates in Submitted stage (have CandidateSubmission records)."""

    allow_user(current_user)

    

    from sqlalchemy import distinct

    

    # Get unique candidates with submission records

    candidate_ids = db.query(distinct(models.CandidateSubmission.candidate_id)).all()

    candidate_ids = [c[0] for c in candidate_ids]

    

    if not candidate_ids:

        return []

    

    query = db.query(models.Candidate).filter(

        models.Candidate.id.in_(candidate_ids),

        models.Candidate.merged_into_id.is_(None)

    ).order_by(models.Candidate.created_at.desc())

    

    candidates = query.all()

    return [normalize_candidate(c) for c in candidates]





@router.get("/stage/interview", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def get_interview_candidates(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """Get all candidates in Interview stage (have scheduled interviews)."""

    allow_user(current_user)

    

    from sqlalchemy import distinct

    

    # Get unique candidates with interview records

    candidate_ids = db.query(distinct(models.CandidateSubmission.candidate_id)).join(

        models.Interview,

        models.Interview.submission_id == models.CandidateSubmission.id

    ).all()

    candidate_ids = [c[0] for c in candidate_ids]

    

    if not candidate_ids:

        return []

    

    query = db.query(models.Candidate).filter(

        models.Candidate.id.in_(candidate_ids),

        models.Candidate.merged_into_id.is_(None)

    ).order_by(models.Candidate.created_at.desc())

    

    candidates = query.all()

    return [normalize_candidate(c) for c in candidates]





@router.get("/stage/offer", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def get_offer_candidates(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """Get all candidates in Offer stage (status='offer')."""

    allow_user(current_user)

    

    query = db.query(models.Candidate).filter(

        models.Candidate.status == "offer",

        models.Candidate.merged_into_id.is_(None)

    ).order_by(models.Candidate.created_at.desc())

    

    candidates = query.all()

    return [normalize_candidate(c) for c in candidates]





@router.get("/stage/joined", response_model=List[schemas.CandidateResponse])

@require_permission("candidates", "view")

async def get_joined_candidates(

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """Get all candidates in Joined stage (status='joined')."""

    allow_user(current_user)

    

    query = db.query(models.Candidate).filter(

        models.Candidate.status == "joined",

        models.Candidate.merged_into_id.is_(None)

    ).order_by(models.Candidate.created_at.desc())

    

    candidates = query.all()

    return [normalize_candidate(c) for c in candidates]





# ============================================================

# GET SINGLE CANDIDATE

@router.get("/{candidate_id}", response_model=schemas.CandidateResponse)

@require_permission("candidates", "view")

async def get_candidate(

    candidate_id: str,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    candidate = db.query(models.Candidate).filter(

        models.Candidate.id == candidate_id

    ).first()



    if not candidate:

        raise HTTPException(404, "Candidate not found")



    # 🔁 Redirect merged → primary

    if candidate.merged_into_id:

        candidate = db.query(models.Candidate).filter(

            models.Candidate.id == candidate.merged_into_id

        ).first()



    resolved_mapping = resolve_candidate_job_account_manager(db, candidate)
    if resolved_mapping:
        if resolved_mapping["job_id"] and not candidate.applied_job_id:
            candidate.applied_job_id = resolved_mapping["job_id"]
        candidate.job_title = resolved_mapping["job_title"] or getattr(
            candidate, "job_title", None
        )
        candidate.account_manager_id = resolved_mapping["account_manager_id"]
        candidate.account_manager_name = resolved_mapping["account_manager_name"]
        candidate.account_manager_email = resolved_mapping["account_manager_email"]

    return normalize_candidate(candidate)





# ============================================================

# RESUME VERSION HISTORY

# ============================================================

@router.get("/{candidate_id}/resume/versions")

@require_permission("candidates", "view")

async def admin_resume_versions(

    candidate_id: str,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):



    allow_user(current_user)



    c = db.query(models.Candidate).filter(models.Candidate.id == candidate_id).first()

    if not c:

        raise HTTPException(404, "Candidate not found")



    return c.resume_versions or []





# ============================================================

# RESTORE RESUME VERSION

# ============================================================

@router.post("/{candidate_id}/resume/restore/{version_id}")

@require_permission("candidates", "update")

async def restore_resume(

    candidate_id: str,

    version_id: str,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):



    allow_user(current_user)



    c = db.query(models.Candidate).filter(models.Candidate.id == candidate_id).first()

    if not c:

        raise HTTPException(404, "Candidate not found")



    match = [v for v in (c.resume_versions or []) if v["version_id"] == version_id]

    if not match:

        raise HTTPException(404, "Version not found")



    c.resume_url = match[0]["url"]

    c.parsed_resume = match[0]["parsed_snapshot"]

    c.last_resume_update = datetime.utcnow()



    db.commit()



    return {"message": "Version restored"}



# ============================================================

# FORWARD PROFILE TO ANOTHER USER

# ============================================================

@router.post("/{candidate_id}/forward")

@require_permission("candidates", "update")

async def forward_profile(

    candidate_id: str,

    data: schemas.ForwardProfileRequest,
    job_id: Optional[str] = Query(None),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):



    allow_user(current_user)



    c = db.query(models.Candidate).filter(models.Candidate.id == candidate_id).first()

    if not c:

        raise HTTPException(404, "Candidate not found")
    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current_user,
        candidate_id=candidate_id,
        job_ref=job_id,
        action="forward_profile",
    )



    user = db.query(models.User).filter(models.User.id == data.forwarded_to_user_id).first()

    if not user:

        raise HTTPException(404, "Target user not found")



    c.forwarded_to = user.username

    c.forward_note = data.note

    c.forwarded_at = datetime.utcnow()

    c.status = "screened"



    db.commit()



    return {"message": "Profile forwarded successfully"}





# ============================================================

# UPDATE CANDIDATE STATUS

# ============================================================

@router.put("/{candidate_id}/status")

@require_permission("candidates", "update")

async def update_status(

    candidate_id: str,

    job_id: Optional[str] = Form(None),

    status: str = Form(...),

    note: Optional[str] = Form(None),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):



    allow_user(current_user)



    c = db.query(models.Candidate).filter(models.Candidate.id == candidate_id).first()

    if not c:

        raise HTTPException(404, "Candidate not found")
    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current_user,
        candidate_id=candidate_id,
        job_ref=job_id,
        action="update_status",
    )



    c.status = status



    timeline = models.CandidateTimeline(

        candidate_id=candidate_id,

        status=status,

        note=note,

        user_id=current_user["id"]

    )



    db.add(timeline)

    db.commit()



    return {"message": "Status updated successfully"}





# ============================================================

# CANDIDATE STATUS TIMELINE

# ============================================================

@router.get("/{candidate_id}/timeline")

@require_permission("candidates", "view")

async def get_timeline(

    candidate_id: str,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):



    allow_user(current_user)



    events = (

        db.query(models.CandidateTimeline)

        .filter(models.CandidateTimeline.candidate_id == candidate_id)

        .order_by(models.CandidateTimeline.created_at.desc())

        .all()

    )

    
    # Build timeline with user role info
    timeline_items = []
    for e in events:
        user_name = ""
        user_role = ""
        if e.user:
            user_name = e.user.full_name or e.user.username or e.user.email or ""
            if e.user.role:
                role_str = e.user.role.value if hasattr(e.user.role, 'value') else str(e.user.role)
                role_str = role_str.lower()
                if 'account' in role_str or 'am' in role_str:
                    user_role = "Account Manager"
                elif 'recruiter' in role_str:
                    user_role = "Recruiter"
                elif 'admin' in role_str:
                    user_role = "Admin"
                elif 'client' in role_str:
                    user_role = "Client"
                else:
                    user_role = role_str.replace('_', ' ').title()
        
        timeline_items.append({
            "id": e.id,
            "status": e.status,
            "note": e.note,
            "by": user_name,
            "role": user_role,
            "at": str(e.created_at) if e.created_at else None,
            "created_at": e.created_at
        })
    
    return timeline_items





# ============================================================

# ADD ADMIN NOTE TO CANDIDATE

# ============================================================

@router.post("/{candidate_id}/notes")

def add_candidate_note(

    candidate_id: str,

    payload: dict,
    job_id: Optional[str] = Query(None),

    db: Session = Depends(get_db),

    current=Depends(get_current_user)

):

    from app.models import Candidate, CandidateNote, Notification



    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()

    if not candidate:

        raise HTTPException(status_code=404, detail="Candidate not found")
    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current,
        candidate_id=candidate_id,
        job_ref=job_id,
        action="add_candidate_note",
    )



    note_text = payload.get("note")

    if not note_text:

        raise HTTPException(status_code=400, detail="Note is required")



    note = CandidateNote(

        candidate_id=candidate_id,

        note=note_text,

        author_id=current.get("id")

    )



    db.add(note)

    db.commit()

    db.refresh(note)



    # 🔥 Candidate Notification Create

    notification = Notification(

        candidate_id=candidate_id,

        title="New Update From Admin",

        message=note_text,

        type="admin_note"

    )



    db.add(notification)

    db.commit()



    return {

        "message": "Note added successfully and notification sent",

        "note_id": note.id

    }







# ============================================================

# SEND BULK EMAIL TO MULTIPLE CANDIDATES

# ============================================================

@router.post("/email/send")

@require_permission("candidates", "update")

async def bulk_email(

    subject: str = Form(...),

    message_body: str = Form(...),

    candidate_ids: List[str] = Form(...),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):



    allow_user(current_user)



    candidates = db.query(models.Candidate).filter(models.Candidate.id.in_(candidate_ids)).all()

    if not candidates:

        raise HTTPException(404, "No candidates found")



    sent = []



    for c in candidates:

        logs = c.email_logs or []

        logs.append({

            "subject": subject,

            "body": message_body,

            "sent_at": datetime.utcnow().isoformat(),

            "sent_by": current_user["id"]

        })

        c.email_logs = logs

        sent.append(c.email)



    db.commit()



    return {

        "message": "Emails successfully sent",

        "sent_to": sent

    }





@router.get("/{candidate_id}/resume/download")

@require_permission("candidates", "view")

async def download_resume(

    candidate_id: str,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):

    allow_user(current_user)



    candidate = (

        db.query(models.Candidate)

        .filter(models.Candidate.id == candidate_id)

        .first()

    )



    if not candidate:

        raise HTTPException(status_code=404, detail="Candidate not found")



    if not candidate.resume_url:

        raise HTTPException(status_code=404, detail="Resume not uploaded")



    file_path = candidate.resume_url



    if not os.path.exists(file_path):

        raise HTTPException(status_code=404, detail="Resume file missing on server")



    filename = os.path.basename(file_path)



    return FileResponse(

        path=file_path,

        filename=filename,

        media_type="application/octet-stream"

    )





# ============================================================

# MERGE MULTIPLE CANDIDATES

# ============================================================

# ============================================================

# MERGE CANDIDATES

@router.post("/merge")

@require_permission("candidates", "update")

async def merge_candidates(

    data: schemas.CandidateMergeRequest,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    allow_user(current_user)



    if not data.candidate_ids or len(data.candidate_ids) < 2:

        raise HTTPException(400, "At least 2 candidates required")



    candidates = (

        db.query(models.Candidate)

        .filter(

            models.Candidate.id.in_(data.candidate_ids),

            models.Candidate.merged_into_id.is_(None)

        )

        .order_by(models.Candidate.created_at.desc())

        .all()

    )



    if len(candidates) < 2:

        raise HTTPException(404, "Candidates not found")



    # ✅ PRIMARY = most recently updated

    primary = candidates[0]

    duplicates = candidates[1:]



    for dup in duplicates:

        # 🔥 Always keep latest real data

        if dup.phone:

            primary.phone = dup.phone

        if dup.resume_url:

            primary.resume_url = dup.resume_url

        if dup.skills:

            primary.skills = dup.skills

        if dup.education:

            primary.education = dup.education

        if dup.experience_years:

            primary.experience_years = dup.experience_years



        # 🔁 Move relations

        for app in dup.applications:

            app.candidate_id = primary.id

        for note in dup.notes:

            note.candidate_id = primary.id

        for event in dup.timeline_events:

            event.candidate_id = primary.id



        # 🔒 Lock duplicate

        dup.status = "merged"

        dup.merged_into_id = primary.id



    db.commit()



    return {

        "message": "Candidates merged successfully",

        "primary_candidate_id": primary.id

    }







@router.patch("/{candidate_id}/classification")

@router.put("/{candidate_id}/classification")

@require_permission("candidates", "update")

async def update_candidate_classification(

    candidate_id: str,

    data: schemas.CandidateClassificationUpdate,
    job_id: Optional[str] = Query(None),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):

    allow_user(current_user)



    candidate = db.query(models.Candidate).filter(

        models.Candidate.id == candidate_id

    ).first()



    if not candidate:

        raise HTTPException(404, "Candidate not found")
    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current_user,
        candidate_id=candidate_id,
        job_ref=job_id,
        action="update_classification",
    )



    if candidate.merged_into_id:

        raise HTTPException(

            status_code=400,

            detail="Merged candidate cannot be modified"

        )



    candidate.classification = data.classification

    db.commit()



    return {

        "message": "Classification updated",

        "classification": candidate.classification

    }



@router.put("/{candidate_id}/verify")

@require_permission("candidates", "update")

def verify_candidate(

    candidate_id: str,
    job_id: Optional[str] = Query(None),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    candidate = db.query(models.Candidate).filter(

        models.Candidate.id == candidate_id

    ).first()



    if not candidate:

        raise HTTPException(404, "Candidate not found")
    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current_user,
        candidate_id=candidate_id,
        job_ref=job_id,
        action="verify_candidate",
    )



    if candidate.merged_into_id:

        raise HTTPException(

            status_code=400,

            detail="Merged candidate cannot be verified"

        )



    if candidate.status in ["verified", "converted"]:

        raise HTTPException(

            status_code=400,

            detail=f"Candidate already {candidate.status}"

        )



    candidate.status = "verified"



    timeline = models.CandidateTimeline(

        candidate_id=candidate.id,

        status="verified",

        note="Verified by recruiter",

        user_id=current_user["id"]

    )



    db.add(timeline)

    db.commit()



    return {

        "message": "Candidate verified successfully",

        "candidateId": candidate.id,

        "status": candidate.status

    }



@router.put("/{candidate_id}/status")

@require_permission("candidates", "update")

async def update_candidate_status(

    candidate_id: str,

    job_id: Optional[str] = Form(None),

    status: str = Form(...),

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user)

):

    allow_user(current_user)



    candidate = db.query(models.Candidate).filter(

        models.Candidate.id == candidate_id

    ).first()



    if not candidate:

        raise HTTPException(404, "Candidate not found")
    _enforce_recruiter_candidate_job_lock(
        db,
        current_user=current_user,
        candidate_id=candidate_id,
        job_ref=job_id,
        action="update_candidate_status",
    )



    if candidate.merged_into_id:

        raise HTTPException(

            status_code=400,

            detail="Merged candidate is read-only"

        )



    if status != "verified":

        raise HTTPException(

            status_code=400,

            detail="Only 'verified' status is allowed"

        )



    candidate.status = "verified"



    timeline = models.CandidateTimeline(

        candidate_id=candidate.id,

        status="verified",

        note="Verified by recruiter",

        user_id=current_user["id"]

    )



    db.add(timeline)

    db.commit()



    return {

        "message": "Candidate verified successfully",

        "candidateId": candidate.id,

        "status": candidate.status

    }


# =============================
# Single Resume Upload (frontend drag/drop)
# =============================
def normalize_resume_identity(
    email: Optional[str],
    phone: Optional[str],
    filename: str,
):
    """
    Enforce DB identity constraints for candidates:
    - At least one of email/phone is required (chk_candidate_identity)
    - Email must be Gmail if present (chk_only_gmail_allowed)
    Returns: (normalized_email, normalized_phone, generated_contact)
    """
    normalized_email = (email or "").strip().lower() or None
    normalized_phone = (phone or "").strip() or None
    generated_contact = False

    if normalized_email and not normalized_email.endswith("@gmail.com"):
        normalized_email = None

    if not normalized_email and not normalized_phone:
        base = os.path.splitext(filename or "resume")[0]
        slug = re.sub(r"[^a-z0-9]+", "", base.lower()).strip()
        if not slug:
            slug = "candidate"
        slug = slug[:24]
        normalized_email = f"{slug}.{uuid.uuid4().hex[:8]}@gmail.com"
        generated_contact = True

    return normalized_email, normalized_phone, generated_contact


def compact_resume_upload_error(error: Exception) -> str:
    message = str(error or "")
    lowered = message.lower()
    if "chk_candidate_identity" in lowered:
        return "Missing email and phone in resume"
    if "chk_only_gmail_allowed" in lowered:
        return "Only Gmail addresses are allowed"
    if "duplicate key value violates unique constraint" in lowered:
        return "Duplicate candidate email already exists"
    if not message:
        return "Failed to import resume"
    return message[:120]


def _pick_first_non_empty(*values):
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _clean_designation_text(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    if len(text) > 120:
        return ""
    low = text.lower()
    if "@" in low or "http://" in low or "https://" in low:
        return ""
    if re.search(r"\b(19|20)\d{2}\b", low):
        return ""
    words = [w for w in re.split(r"[\s,/|()\-]+", low) if w]
    if len(words) > 8:
        return ""
    noise_words = {
        "worked",
        "working",
        "experience",
        "knowledge",
        "responsible",
        "responsibilities",
        "project",
        "projects",
        "ticket",
        "tickets",
        "clarification",
        "issues",
        "resolution",
        "candidate",
        "college",
        "university",
        "payroll",
        "involved",
        "support",
    }
    if any(word in noise_words for word in words):
        return ""
    role_words = {
        "engineer",
        "developer",
        "manager",
        "analyst",
        "consultant",
        "architect",
        "lead",
        "specialist",
        "executive",
        "officer",
        "intern",
        "designer",
        "administrator",
        "devops",
        "qa",
        "tester",
        "associate",
        "director",
        "head",
        "coordinator",
        "principal",
        "staff",
        "recruiter",
        "accountant",
        "moderator",
        "moderation",
    }
    if not any(word in role_words for word in words):
        return ""
    return text


def _clean_location_text(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    if len(text) > 60:
        return ""
    low = text.lower()
    if "@" in low or "http://" in low or "https://" in low:
        return ""
    words = [w for w in re.split(r"[\s,]+", low) if w]
    if len(words) > 6:
        return ""
    noise_words = {
        "worked",
        "working",
        "experience",
        "knowledge",
        "responsible",
        "responsibilities",
        "project",
        "projects",
        "ticket",
        "tickets",
        "clarification",
        "issues",
        "resolution",
        "candidate",
        "college",
        "university",
        "payroll",
        "involved",
        "support",
    }
    if any(word in noise_words for word in words):
        return ""
    if sum(1 for ch in text if ch.isdigit()) > 3:
        return ""
    return text


def _extract_resume_profile_fields(data: dict):
    designation = _clean_designation_text(
        _pick_first_non_empty(
            data.get("current_designation"),
            data.get("current_role"),
            data.get("designation"),
            data.get("designation_title"),
            data.get("title"),
            data.get("role"),
            (data.get("work_history") or [{}])[0].get("designation")
            if isinstance(data.get("work_history"), list) and data.get("work_history")
            else None,
            (data.get("work_history") or [{}])[0].get("role")
            if isinstance(data.get("work_history"), list) and data.get("work_history")
            else None,
            (data.get("work_history") or [{}])[0].get("title")
            if isinstance(data.get("work_history"), list) and data.get("work_history")
            else None,
            (data.get("work_experience") or [{}])[0].get("designation")
            if isinstance(data.get("work_experience"), list) and data.get("work_experience")
            else None,
            (data.get("work_experience") or [{}])[0].get("role")
            if isinstance(data.get("work_experience"), list) and data.get("work_experience")
            else None,
            (data.get("work_experience") or [{}])[0].get("title")
            if isinstance(data.get("work_experience"), list) and data.get("work_experience")
            else None,
        )
    )
    company = _pick_first_non_empty(
        data.get("current_company"),
        data.get("current_employer"),
        data.get("company"),
        (data.get("work_history") or [{}])[0].get("company")
        if isinstance(data.get("work_history"), list) and data.get("work_history")
        else None,
        (data.get("work_experience") or [{}])[0].get("company")
        if isinstance(data.get("work_experience"), list) and data.get("work_experience")
        else None,
    )
    current_location = _clean_location_text(
        _pick_first_non_empty(
            data.get("current_location"),
            data.get("location"),
            data.get("city"),
            data.get("preferred_location"),
        )
    )
    city = _clean_location_text(_pick_first_non_empty(data.get("city"), current_location))
    return designation, company, current_location, city


def _to_bool_or_none(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"yes", "y", "true", "1"}:
        return True
    if text in {"no", "n", "false", "0"}:
        return False
    return None


def _to_float_or_none(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text.replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group(1))
    except Exception:
        return None


def _to_notice_days(value):
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    qty = int(match.group(1))
    if "month" in text:
        return qty * 30
    if "week" in text:
        return qty * 7
    return qty


def _apply_parsed_profile_fields_to_candidate(
    candidate: models.Candidate,
    data: dict,
    *,
    parsed_designation: str,
    parsed_company: str,
    parsed_location: str,
    parsed_city: str,
) -> None:
    if not isinstance(data, dict):
        return

    candidate.current_job_title = parsed_designation or getattr(candidate, "current_job_title", "")
    candidate.current_role = parsed_designation or getattr(candidate, "current_role", "")
    candidate.current_employer = parsed_company or getattr(candidate, "current_employer", "")
    candidate.current_location = parsed_location or getattr(candidate, "current_location", "")
    candidate.city = parsed_city or getattr(candidate, "city", "")

    candidate.gender = _pick_first_non_empty(data.get("gender"), getattr(candidate, "gender", None))
    candidate.date_of_birth = _pick_first_non_empty(
        data.get("date_of_birth"),
        data.get("dob"),
        getattr(candidate, "date_of_birth", None),
    )

    notice_period = _pick_first_non_empty(data.get("notice_period"), getattr(candidate, "notice_period", None))
    if notice_period:
        candidate.notice_period = notice_period
    notice_days = data.get("notice_period_days")
    if notice_days is None:
        notice_days = _to_notice_days(notice_period)
    try:
        if notice_days is not None:
            candidate.notice_period_days = int(notice_days)
    except Exception:
        pass

    current_ctc = _to_float_or_none(data.get("current_ctc"))
    if current_ctc is not None:
        candidate.current_ctc = current_ctc
    expected_ctc = _pick_first_non_empty(data.get("expected_ctc"), getattr(candidate, "expected_ctc", None))
    if expected_ctc:
        candidate.expected_ctc = expected_ctc

    candidate.current_address = _pick_first_non_empty(
        data.get("current_address"),
        getattr(candidate, "current_address", None),
    )
    candidate.permanent_address = _pick_first_non_empty(
        data.get("permanent_address"),
        getattr(candidate, "permanent_address", None),
    )
    candidate.pincode = _pick_first_non_empty(data.get("pincode"), getattr(candidate, "pincode", None))
    candidate.preferred_location = _pick_first_non_empty(
        data.get("preferred_location"),
        getattr(candidate, "preferred_location", None),
    )

    relocate = _to_bool_or_none(
        _pick_first_non_empty(
            data.get("ready_to_relocate"),
            data.get("willing_to_relocate"),
        )
    )
    if relocate is not None:
        candidate.willing_to_relocate = relocate
        candidate.ready_to_relocate = "Yes" if relocate else "No"

    languages = data.get("languages")
    if isinstance(languages, list):
        candidate.languages_known = languages

    work_history = data.get("work_history")
    if not isinstance(work_history, list):
        work_history = data.get("work_experience")
    if isinstance(work_history, list):
        candidate.work_history = work_history

    education_history = data.get("education_history")
    if isinstance(education_history, list):
        candidate.education_history = education_history

    projects = data.get("projects")
    if isinstance(projects, list):
        candidate.projects = projects

    certifications_text = _pick_first_non_empty(data.get("certifications_text"), getattr(candidate, "certifications_text", None))
    if certifications_text:
        candidate.certifications_text = certifications_text


@router.post('/upload-resume')
@require_permission('candidates', 'create')
async def upload_resume(
    file: UploadFile = File(...),
    duplicate_option: str = Form('overwrite'),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    allow_user(current_user)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    original_name = file.filename or 'resume.pdf'
    safe_name = f"{uuid.uuid4().hex}_{original_name}"
    path = os.path.join(UPLOAD_DIR, safe_name)
    with open(path, 'wb') as f:
        f.write(await file.read())

    parsed = parse_resume_file(path) or {}
    data = parsed.get('data') if isinstance(parsed, dict) else parsed
    if not isinstance(data, dict):
        raise HTTPException(400, 'Failed to parse resume')

    raw_email = (data.get('email') or '').strip() or None
    raw_phone = (data.get('phone') or '').strip() or None
    email, phone, generated_contact = normalize_resume_identity(
        raw_email,
        raw_phone,
        original_name,
    )
    if generated_contact:
        data['identity_note'] = 'Generated temporary Gmail due to missing email/phone in resume'
    elif raw_email and raw_email.lower() != (email or "").lower():
        data['identity_note'] = 'Converted non-Gmail email to system-compliant contact'
    parsed_designation, parsed_company, parsed_location, parsed_city = _extract_resume_profile_fields(data)
    full_name = (data.get('full_name') or '').strip()
    if not full_name:
        base = os.path.splitext(original_name)[0]
        base = base.replace('_', ' ').replace('-', ' ')
        full_name = base.title() if base else 'Unknown'

    candidate = None
    if email:
        candidate = db.query(models.Candidate).filter(func.lower(models.Candidate.email) == email.lower()).first()
    if not candidate and phone:
        candidate = db.query(models.Candidate).filter(models.Candidate.phone == phone).first()

    status = 'created'
    if candidate:
        if duplicate_option == 'overwrite':
            candidate.full_name = full_name or candidate.full_name
            candidate.email = email or candidate.email
            candidate.phone = phone or candidate.phone
            candidate.skills = data.get('skills') or candidate.skills
            candidate.education = data.get('education') or candidate.education
            candidate.experience_years = data.get('experience_years') or candidate.experience_years
            _apply_parsed_profile_fields_to_candidate(
                candidate,
                data,
                parsed_designation=parsed_designation,
                parsed_company=parsed_company,
                parsed_location=parsed_location,
                parsed_city=parsed_city,
            )
            candidate.resume_url = f"/uploads/resumes/{safe_name}"
            candidate.parsed_resume = parsed
            candidate.parsed_data_json = data
            candidate.updated_at = datetime.utcnow()
            if str(candidate.status).lower() in {"new", "verified"}:
                candidate.status = models.CandidateStatus.sourced
            if not candidate.source:
                candidate.source = "Resume Upload"
            status = 'updated'
        else:
            status = 'duplicate'
    else:
        candidate = models.Candidate(
            public_id=models.generate_candidate_public_id_from_org(db),
            full_name=full_name or 'Unknown',
            email=email,
            phone=phone,
            skills=data.get('skills') or [],
            education=data.get('education') or [],
            experience_years=data.get('experience_years') or 0,
            current_job_title='',
            current_role='',
            current_employer='',
            current_location='',
            city='',
            resume_url=f"/uploads/resumes/{safe_name}",
            status=models.CandidateStatus.sourced,
            source='Resume Upload',
            parsed_resume=parsed,
            parsed_data_json=data,
            created_at=datetime.utcnow(),
        )
        _apply_parsed_profile_fields_to_candidate(
            candidate,
            data,
            parsed_designation=parsed_designation,
            parsed_company=parsed_company,
            parsed_location=parsed_location,
            parsed_city=parsed_city,
        )
        db.add(candidate)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(400, compact_resume_upload_error(e))
    return {'message': 'success', 'status': status, 'candidate_id': getattr(candidate, 'public_id', None)}

# =============================
# Bulk Resume Upload (admin single/multi)
# =============================
@router.post('/bulk-resume-upload')
@require_permission('candidates', 'create')
async def bulk_resume_upload(
    files: List[UploadFile] = File(...),
    duplicate_option: str = Form('overwrite'),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    allow_user(current_user)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    results = []
    success = failed = duplicates = updated = 0

    def derive_name_from_filename(name: str) -> str:
        base = os.path.splitext(name or '')[0]
        base = base.replace('_', ' ').replace('-', ' ')
        base = re.sub(r"\d+", " ", base)
        base = re.sub(r"\s+", " ", base).strip()
        return base.title() if base else "Unknown"

    for file in files:
        original_name = file.filename or 'resume.pdf'
        safe_name = f"{uuid.uuid4().hex}_{original_name}"
        path = os.path.join(UPLOAD_DIR, safe_name)
        try:
            with open(path, 'wb') as f:
                f.write(await file.read())
            parsed = parse_resume_file(path) or {}
            data = parsed.get('data') if isinstance(parsed, dict) else parsed
            if not isinstance(data, dict):
                raise ValueError('parse failed')
            raw_email = (data.get('email') or '').strip() or None
            raw_phone = (data.get('phone') or '').strip() or None
            email, phone, generated_contact = normalize_resume_identity(
                raw_email,
                raw_phone,
                original_name,
            )
            if generated_contact:
                data['identity_note'] = 'Generated temporary Gmail due to missing email/phone in resume'
            elif raw_email and raw_email.lower() != (email or "").lower():
                data['identity_note'] = 'Converted non-Gmail email to system-compliant contact'
            parsed_designation, parsed_company, parsed_location, parsed_city = _extract_resume_profile_fields(data)
            full_name = (data.get('full_name') or '').strip()
            if not full_name:
                full_name = derive_name_from_filename(original_name)

            candidate = None
            if email:
                candidate = db.query(models.Candidate).filter(func.lower(models.Candidate.email)==email.lower()).first()
            if not candidate and phone:
                candidate = db.query(models.Candidate).filter(models.Candidate.phone==phone).first()

            status = 'created'
            if candidate:
                if duplicate_option == 'overwrite':
                    candidate.full_name = full_name or candidate.full_name
                    candidate.email = email or candidate.email
                    candidate.phone = phone or candidate.phone
                    candidate.skills = data.get('skills') or candidate.skills
                    candidate.education = data.get('education') or candidate.education
                    candidate.experience_years = data.get('experience_years') or candidate.experience_years
                    _apply_parsed_profile_fields_to_candidate(
                        candidate,
                        data,
                        parsed_designation=parsed_designation,
                        parsed_company=parsed_company,
                        parsed_location=parsed_location,
                        parsed_city=parsed_city,
                    )
                    candidate.resume_url = f"/uploads/resumes/{safe_name}"
                    candidate.parsed_resume = parsed
                    candidate.parsed_data_json = data
                    candidate.updated_at = datetime.utcnow()
                    if str(candidate.status).lower() in {"new", "verified"}:
                        candidate.status = models.CandidateStatus.sourced
                    if not candidate.source:
                        candidate.source = "Bulk Resume Upload"
                    status = 'updated'
                else:
                    status = 'duplicate'; duplicates +=1
            else:
                candidate = models.Candidate(
                    public_id=models.generate_candidate_public_id_from_org(db),
                    full_name=full_name or 'Unknown',
                    email=email,
                    phone=phone,
                    skills=data.get('skills') or [],
                    education=data.get('education') or [],
                    experience_years=data.get('experience_years') or 0,
                    current_job_title='',
                    current_role='',
                    current_employer='',
                    current_location='',
                    city='',
                    resume_url=f"/uploads/resumes/{safe_name}",
                    status=models.CandidateStatus.sourced,
                    source='Bulk Resume Upload',
                    parsed_resume=parsed,
                    parsed_data_json=data,
                    created_at=datetime.utcnow(),
                )
                _apply_parsed_profile_fields_to_candidate(
                    candidate,
                    data,
                    parsed_designation=parsed_designation,
                    parsed_company=parsed_company,
                    parsed_location=parsed_location,
                    parsed_city=parsed_city,
                )
                db.add(candidate)
            db.commit()
            if status == 'updated':
                updated += 1
                success += 1
            elif status == 'created':
                success += 1
            results.append({
                'resume': original_name,
                'candidate_id': getattr(candidate,'public_id', None),
                'name': full_name or 'Unknown',
                'email': email,
                'status': status
            })
        except Exception as e:
            db.rollback()
            failed +=1
            results.append({
                'resume': original_name,
                'candidate_id': None,
                'name': 'Unknown',
                'email': email if 'email' in locals() else None,
                'status': f'Failed - {compact_resume_upload_error(e)[:60]}'
            })
    total = len(files)
    return {
        'total_processed': total,
        'success': success,
        'failed': failed,
        'duplicates': duplicates,
        'updated': updated,
        'results': results,
    }


# Helper utilities for name handling

def normalize_candidate_name(value: str) -> str:
    return str(value or '').strip()


def is_likely_bad_candidate_name(value: Optional[str]) -> bool:
    if not value:
        return True
    cleaned = normalize_candidate_name(value)
    if not cleaned:
        return True
    lower = cleaned.lower()
    if any(ch.isdigit() for ch in cleaned):
        return True
    bad_tokens = {
        'resume','curriculum','vitae','cv','profile','summary','developer','engineer','analyst','consultant','manager',
        'architect','programmer','specialist','intern','lead','associate','designer','tester','administrator','devops',
        'fullstack','frontend','backend','software','data','scientist','design','development','linux','sap','abap',
        'skill','skills','professional','technical','expertise','experience','component','project','projects','portfolio',
        'objective','generator','story','comic','book','capstone'
    }
    words = lower.split()
    if len(words) > 5:
        return True
    if any(w in bad_tokens for w in words):
        return True
    return False


def derive_candidate_name_from_email(email: str) -> Optional[str]:
    if not email or '@' not in email:
        return None
    local = email.split('@',1)[0]
    local = re.sub(r"\d+", " ", local)
    local = re.sub(r"[._-]+", " ", local)
    local = re.sub(r"\s+", " ", local).strip()
    if not local:
        return None
    parts = local.split()[:4]
    return ' '.join(p.capitalize() for p in parts) if parts else None


def derive_candidate_name_from_filename(filename: str) -> str:
    base = os.path.splitext(filename or '')[0]
    base = base.replace('_',' ').replace('-',' ')
    base = re.sub(r"\d+", " ", base)
    base = re.sub(r"\s+", " ", base).strip()
    return base.title() if base else 'Unknown'


# Helper utilities for name handling

def normalize_candidate_name(value: str) -> str:
    return str(value or '').strip()


def is_likely_bad_candidate_name(value: Optional[str]) -> bool:
    if not value:
        return True
    cleaned = normalize_candidate_name(value)
    if not cleaned:
        return True
    lower = cleaned.lower()
    if any(ch.isdigit() for ch in cleaned):
        return True
    bad_tokens = {
        'resume','curriculum','vitae','cv','profile','summary','developer','engineer','analyst','consultant','manager',
        'architect','programmer','specialist','intern','lead','associate','designer','tester','administrator','devops',
        'fullstack','frontend','backend','software','data','scientist','design','development','linux','sap','abap',
        'skill','skills','professional','technical','expertise','experience','component','project'
    }
    words = lower.split()
    if len(words) > 5:
        return True
    if any(w in bad_tokens for w in words):
        return True
    return False


def derive_candidate_name_from_email(email: str) -> Optional[str]:
    if not email or '@' not in email:
        return None
    local = email.split('@',1)[0]
    local = re.sub(r"\d+", " ", local)
    local = re.sub(r"[._-]+", " ", local)
    local = re.sub(r"\s+", " ", local).strip()
    if not local:
        return None
    parts = local.split()[:4]
    return ' '.join(p.capitalize() for p in parts) if parts else None


def derive_candidate_name_from_filename(filename: str) -> str:
    base = os.path.splitext(filename or '')[0]
    base = base.replace('_',' ').replace('-',' ')
    base = re.sub(r"\d+", " ", base)
    base = re.sub(r"\s+", " ", base).strip()
    return base.title() if base else 'Unknown'

# Single resume upload (used by frontend admin upload)
@router.post('/upload-resume')
@require_permission('candidates', 'create')
async def upload_resume(
    file: UploadFile = File(...),
    duplicate_option: str = Form('overwrite'),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    allow_user(current_user)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    original_name = file.filename or 'resume.pdf'
    safe_name = f"{uuid.uuid4().hex}_{original_name}"
    path = os.path.join(UPLOAD_DIR, safe_name)
    with open(path, 'wb') as f:
        f.write(await file.read())

    parsed = parse_resume_file(path) or {}
    data = parsed.get('data') if isinstance(parsed, dict) else parsed
    if not isinstance(data, dict):
        raise HTTPException(400, 'Failed to parse resume')

    raw_email = (data.get('email') or '').strip() or None
    raw_phone = (data.get('phone') or '').strip() or None
    email, phone, generated_contact = normalize_resume_identity(
        raw_email,
        raw_phone,
        original_name,
    )
    if generated_contact:
        data['identity_note'] = 'Generated temporary Gmail due to missing email/phone in resume'
    elif raw_email and raw_email.lower() != (email or "").lower():
        data['identity_note'] = 'Converted non-Gmail email to system-compliant contact'
    parsed_designation, parsed_company, parsed_location, parsed_city = _extract_resume_profile_fields(data)
    parsed_name = (data.get('full_name') or '').strip()
    email_name = derive_candidate_name_from_email(email) if email else None
    filename_name = derive_candidate_name_from_filename(original_name)
    full_name = parsed_name
    if is_likely_bad_candidate_name(full_name):
        full_name = email_name or filename_name or 'Unknown'

    candidate = None
    if email:
        candidate = db.query(models.Candidate).filter(func.lower(models.Candidate.email)==email.lower()).first()
    if not candidate and phone:
        candidate = db.query(models.Candidate).filter(models.Candidate.phone==phone).first()

    status = 'created'
    if candidate:
        if duplicate_option == 'overwrite':
            candidate.full_name = full_name or candidate.full_name
            candidate.email = email or candidate.email
            candidate.phone = phone or candidate.phone
            candidate.skills = data.get('skills') or candidate.skills
            candidate.education = data.get('education') or candidate.education
            candidate.experience_years = data.get('experience_years') or candidate.experience_years
            _apply_parsed_profile_fields_to_candidate(
                candidate,
                data,
                parsed_designation=parsed_designation,
                parsed_company=parsed_company,
                parsed_location=parsed_location,
                parsed_city=parsed_city,
            )
            candidate.resume_url = f"/uploads/resumes/{safe_name}"
            candidate.parsed_resume = parsed
            candidate.parsed_data_json = data
            candidate.updated_at = datetime.utcnow()
            if str(candidate.status).lower() in {"new", "verified"}:
                candidate.status = models.CandidateStatus.sourced
            if not candidate.source:
                candidate.source = "Resume Upload"
            status = 'updated'
        else:
            status = 'duplicate'
    else:
        candidate = models.Candidate(
            public_id=models.generate_candidate_public_id_from_org(db),
            full_name=full_name or 'Unknown',
            email=email,
            phone=phone,
            skills=data.get('skills') or [],
            education=data.get('education') or [],
            experience_years=data.get('experience_years') or 0,
            current_job_title='',
            current_role='',
            current_employer='',
            current_location='',
            city='',
            resume_url=f"/uploads/resumes/{safe_name}",
            status=models.CandidateStatus.sourced,
            source='Resume Upload',
            parsed_resume=parsed,
            parsed_data_json=data,
            created_at=datetime.utcnow(),
        )
        _apply_parsed_profile_fields_to_candidate(
            candidate,
            data,
            parsed_designation=parsed_designation,
            parsed_company=parsed_company,
            parsed_location=parsed_location,
            parsed_city=parsed_city,
        )
        db.add(candidate)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(400, compact_resume_upload_error(e))
    return {'message': 'success', 'status': status, 'candidate_id': getattr(candidate,'public_id', None)}


# Bulk resume upload (admin drag/drop single/multi)
@router.post('/bulk-resume-upload')
@require_permission('candidates', 'create')
async def bulk_resume_upload(
    files: List[UploadFile] = File(...),
    duplicate_option: str = Form('overwrite'),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    allow_user(current_user)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    results=[]; success=failed=duplicates=updated=0

    for file in files:
        original_name = file.filename or 'resume.pdf'
        safe_name = f"{uuid.uuid4().hex}_{original_name}"
        path = os.path.join(UPLOAD_DIR, safe_name)
        try:
            with open(path, 'wb') as f:
                f.write(await file.read())
            parsed = parse_resume_file(path) or {}
            data = parsed.get('data') if isinstance(parsed, dict) else parsed
            if not isinstance(data, dict):
                raise ValueError('parse failed')
            raw_email = (data.get('email') or '').strip() or None
            raw_phone = (data.get('phone') or '').strip() or None
            email, phone, generated_contact = normalize_resume_identity(
                raw_email,
                raw_phone,
                original_name,
            )
            if generated_contact:
                data['identity_note'] = 'Generated temporary Gmail due to missing email/phone in resume'
            elif raw_email and raw_email.lower() != (email or "").lower():
                data['identity_note'] = 'Converted non-Gmail email to system-compliant contact'
            parsed_designation, parsed_company, parsed_location, parsed_city = _extract_resume_profile_fields(data)
            parsed_name = (data.get('full_name') or '').strip()
            email_name = derive_candidate_name_from_email(email) if email else None
            filename_name = derive_candidate_name_from_filename(original_name)
            full_name = parsed_name
            if is_likely_bad_candidate_name(full_name):
                full_name = email_name or filename_name or 'Unknown'

            candidate = None
            if email:
                candidate = db.query(models.Candidate).filter(func.lower(models.Candidate.email)==email.lower()).first()
            if not candidate and phone:
                candidate = db.query(models.Candidate).filter(models.Candidate.phone==phone).first()

            status='created'
            if candidate:
                if duplicate_option == 'overwrite':
                    candidate.full_name = full_name or candidate.full_name
                    candidate.email = email or candidate.email
                    candidate.phone = phone or candidate.phone
                    candidate.skills = data.get('skills') or candidate.skills
                    candidate.education = data.get('education') or candidate.education
                    candidate.experience_years = data.get('experience_years') or candidate.experience_years
                    _apply_parsed_profile_fields_to_candidate(
                        candidate,
                        data,
                        parsed_designation=parsed_designation,
                        parsed_company=parsed_company,
                        parsed_location=parsed_location,
                        parsed_city=parsed_city,
                    )
                    candidate.resume_url = f"/uploads/resumes/{safe_name}"
                    candidate.parsed_resume = parsed
                    candidate.parsed_data_json = data
                    candidate.updated_at = datetime.utcnow()
                    if str(candidate.status).lower() in {"new", "verified"}:
                        candidate.status = models.CandidateStatus.sourced
                    if not candidate.source:
                        candidate.source = "Bulk Resume Upload"
                    status='updated'
                else:
                    status='duplicate'; duplicates+=1
            else:
                candidate = models.Candidate(
                    public_id=models.generate_candidate_public_id_from_org(db),
                    full_name=full_name or 'Unknown',
                    email=email,
                    phone=phone,
                    skills=data.get('skills') or [],
                    education=data.get('education') or [],
                    experience_years=data.get('experience_years') or 0,
                    current_job_title='',
                    current_role='',
                    current_employer='',
                    current_location='',
                    city='',
                    resume_url=f"/uploads/resumes/{safe_name}",
                    status=models.CandidateStatus.sourced,
                    source='Bulk Resume Upload',
                    parsed_resume=parsed,
                    parsed_data_json=data,
                    created_at=datetime.utcnow(),
                )
                _apply_parsed_profile_fields_to_candidate(
                    candidate,
                    data,
                    parsed_designation=parsed_designation,
                    parsed_company=parsed_company,
                    parsed_location=parsed_location,
                    parsed_city=parsed_city,
                )
                db.add(candidate)
            db.commit()
            if status == 'updated':
                updated += 1
                success += 1
            elif status == 'created':
                success += 1
            results.append({'resume': original_name,'candidate_id': getattr(candidate,'public_id',None),'name': full_name or 'Unknown','email': email,'status': status})
        except Exception as e:
            db.rollback()
            failed+=1
            results.append({'resume': original_name,'candidate_id': None,'name': 'Unknown','email': email if 'email' in locals() else None,'status': f'Failed - {compact_resume_upload_error(e)[:60]}'})
    total=len(files)
    return {'total_processed': total,'success': success,'failed': failed,'duplicates': duplicates,'updated': updated,'results': results}


async def _process_bulk_resume_upload_async(
    *,
    task_id: str,
    files_payload: List[dict],
    duplicate_option: str,
) -> None:
    db = SessionLocal()
    try:
        task_queue.update_task_progress(task_id, 1, TaskStatus.PROCESSING)
        os.makedirs(UPLOAD_DIR, exist_ok=True)

        results = []
        success = failed = duplicates = updated = 0
        total = len(files_payload)

        for index, payload in enumerate(files_payload, start=1):
            original_name = payload.get("filename") or "resume.pdf"
            safe_name = f"{uuid.uuid4().hex}_{original_name}"
            path = os.path.join(UPLOAD_DIR, safe_name)
            email = None

            try:
                with open(path, "wb") as f:
                    f.write(payload.get("content") or b"")

                parsed = parse_resume_file(path) or {}
                data = parsed.get("data") if isinstance(parsed, dict) else parsed
                if not isinstance(data, dict):
                    raise ValueError("parse failed")

                raw_email = (data.get("email") or "").strip() or None
                raw_phone = (data.get("phone") or "").strip() or None
                email, phone, generated_contact = normalize_resume_identity(
                    raw_email,
                    raw_phone,
                    original_name,
                )
                if generated_contact:
                    data["identity_note"] = "Generated temporary Gmail due to missing email/phone in resume"
                elif raw_email and raw_email.lower() != (email or "").lower():
                    data["identity_note"] = "Converted non-Gmail email to system-compliant contact"

                parsed_designation, parsed_company, parsed_location, parsed_city = _extract_resume_profile_fields(data)
                parsed_name = (data.get("full_name") or "").strip()
                email_name = derive_candidate_name_from_email(email) if email else None
                filename_name = derive_candidate_name_from_filename(original_name)
                full_name = parsed_name
                if is_likely_bad_candidate_name(full_name):
                    full_name = email_name or filename_name or "Unknown"

                candidate = None
                if email:
                    candidate = (
                        db.query(models.Candidate)
                        .filter(func.lower(models.Candidate.email) == email.lower())
                        .first()
                    )
                if not candidate and phone:
                    candidate = db.query(models.Candidate).filter(models.Candidate.phone == phone).first()

                status = "created"
                if candidate:
                    if duplicate_option == "overwrite":
                        candidate.full_name = full_name or candidate.full_name
                        candidate.email = email or candidate.email
                        candidate.phone = phone or candidate.phone
                        candidate.skills = data.get("skills") or candidate.skills
                        candidate.education = data.get("education") or candidate.education
                        candidate.experience_years = data.get("experience_years") or candidate.experience_years
                        _apply_parsed_profile_fields_to_candidate(
                            candidate,
                            data,
                            parsed_designation=parsed_designation,
                            parsed_company=parsed_company,
                            parsed_location=parsed_location,
                            parsed_city=parsed_city,
                        )
                        candidate.resume_url = f"/uploads/resumes/{safe_name}"
                        candidate.parsed_resume = parsed
                        candidate.parsed_data_json = data
                        candidate.updated_at = datetime.utcnow()
                        if str(candidate.status).lower() in {"new", "verified"}:
                            candidate.status = models.CandidateStatus.sourced
                        if not candidate.source:
                            candidate.source = "Bulk Resume Upload"
                        status = "updated"
                    else:
                        status = "duplicate"
                        duplicates += 1
                else:
                    candidate = models.Candidate(
                        public_id=models.generate_candidate_public_id_from_org(db),
                        full_name=full_name or "Unknown",
                        email=email,
                        phone=phone,
                        skills=data.get("skills") or [],
                        education=data.get("education") or [],
                        experience_years=data.get("experience_years") or 0,
                        current_job_title="",
                        current_role="",
                        current_employer="",
                        current_location="",
                        city="",
                        resume_url=f"/uploads/resumes/{safe_name}",
                        status=models.CandidateStatus.sourced,
                        source="Bulk Resume Upload",
                        parsed_resume=parsed,
                        parsed_data_json=data,
                        created_at=datetime.utcnow(),
                    )
                    _apply_parsed_profile_fields_to_candidate(
                        candidate,
                        data,
                        parsed_designation=parsed_designation,
                        parsed_company=parsed_company,
                        parsed_location=parsed_location,
                        parsed_city=parsed_city,
                    )
                    db.add(candidate)

                db.commit()

                if status == "updated":
                    updated += 1
                    success += 1
                elif status == "created":
                    success += 1

                results.append(
                    {
                        "resume": original_name,
                        "candidate_id": getattr(candidate, "public_id", None),
                        "name": full_name or "Unknown",
                        "email": email,
                        "status": status,
                    }
                )
            except Exception as e:
                db.rollback()
                failed += 1
                results.append(
                    {
                        "resume": original_name,
                        "candidate_id": None,
                        "name": "Unknown",
                        "email": email,
                        "status": f"Failed - {compact_resume_upload_error(e)[:60]}",
                    }
                )
            finally:
                progress = int((index / total) * 100) if total else 100
                task_queue.update_task_progress(task_id, progress)
                await asyncio.sleep(0)

        task_queue.complete_task(
            task_id,
            {
                "total_processed": total,
                "success": success,
                "failed": failed,
                "duplicates": duplicates,
                "updated": updated,
                "results": results,
            },
        )
    except Exception as e:
        task_queue.fail_task(task_id, compact_resume_upload_error(e))
    finally:
        db.close()


@router.post("/bulk-resume-upload-async")
@require_permission("candidates", "create")
async def bulk_resume_upload_async(
    files: List[UploadFile] = File(...),
    duplicate_option: str = Form("overwrite"),
    current_user=Depends(get_current_user),
):
    allow_user(current_user)

    if not files:
        raise HTTPException(400, "No files provided")
    if len(files) > 200:
        raise HTTPException(400, "Maximum 200 files allowed per async batch")

    duplicate_option = (duplicate_option or "overwrite").strip().lower()
    if duplicate_option not in {"overwrite", "skip"}:
        raise HTTPException(400, "duplicate_option must be either 'overwrite' or 'skip'")

    files_payload = []
    for file in files:
        content = await file.read()
        if not content:
            continue
        files_payload.append(
            {
                "filename": file.filename or "resume.pdf",
                "content_type": file.content_type,
                "content": content,
            }
        )

    if not files_payload:
        raise HTTPException(400, "No valid files to process")

    task_id = task_queue.create_task(
        "bulk_resume_upload",
        metadata={
            "file_count": len(files_payload),
            "duplicate_option": duplicate_option,
            "requested_by": current_user.get("id"),
        },
    )

    asyncio.create_task(
        _process_bulk_resume_upload_async(
            task_id=task_id,
            files_payload=files_payload,
            duplicate_option=duplicate_option,
        )
    )

    return {
        "status": "pending",
        "task_id": task_id,
        "message": "Bulk resume upload started",
        "total_files": len(files_payload),
    }


@router.get("/bulk-upload-status/{task_id}")
@require_permission("candidates", "view")
def bulk_upload_status(task_id: str, current_user=Depends(get_current_user)):
    allow_user(current_user)

    status = get_task_status(task_id)
    if status.get("error"):
        raise HTTPException(404, status["error"])
    return status
