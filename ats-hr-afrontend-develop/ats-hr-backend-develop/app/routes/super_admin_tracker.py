from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
import csv
import re
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import and_, asc, case, desc, func, or_
from sqlalchemy.orm import Session, aliased

from app import models
from app.auth import get_current_user
from app.db import get_db

router = APIRouter(prefix="/v1/super-admin/tracker", tags=["Super Admin Tracker"])

SHORTLISTED = {"Profile Submitted", "Shortlisted"}
INTERVIEWS = {"Interview Scheduled", "Interview No Show"}
SELECTED = {"Selected", "Offered"}
REJECTED = {"Screen Reject Internal", "Screen Reject Client", "Interview Reject", "Joining No Show"}
PORTAL_SHORTLISTED = {"am_shortlisted", "client_shortlisted", "shortlisted"}
PORTAL_INTERVIEWS = {"interview_scheduled", "interview_done", "interview_completed", "no_show"}
PORTAL_SELECTED = {"selected", "offer_extended", "offer_accepted", "hired", "joined"}
PORTAL_JOINED = {"joined", "hired"}
PORTAL_ON_HOLD = {"hold_revisit", "am_hold", "client_hold", "on_hold", "hold", "paused"}

SECTION_ALIAS = {
    "submission details": "submissions",
    "selection details": "selections",
    "channel partner details": "channel_partners",
    "client invoice details": "client_invoices",
    "channel partner invoice": "cp_invoices",
}

SECTION_MODEL = {
    "submissions": models.TrackerSubmission,
    "selections": models.TrackerSelection,
    "channel_partners": models.TrackerChannelPartner,
    "client_invoices": models.TrackerClientInvoice,
    "cp_invoices": models.TrackerCPInvoice,
}

SUBMISSION_STATUS_GROUPS = {
    "total": None,
    "shortlisted": SHORTLISTED,
    "interviews": INTERVIEWS,
    "selected": SELECTED,
    "joined": {"Joined"},
    "rejected": REJECTED,
    "duplicates": {"Duplicate"},
    "on_hold": {"Req on Hold"},
}


def _must_access(user: Dict[str, Any]) -> None:
    role = str((user or {}).get("role") or "").strip().lower()
    if role not in {"super_admin", "admin"}:
        raise HTTPException(403, "Super Admin/Admin access required")


def _to_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d\.\-]", "", str(v).replace(",", ""))
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _window(period: str, date_from: Optional[str], date_to: Optional[str]) -> Tuple[date, date, date, date]:
    today = datetime.utcnow().date()
    p = (period or "month").lower()
    if p in {"week", "this_week"}:
        s, e = today - timedelta(days=today.weekday()), today
    elif p in {"today"}:
        s = e = today
    elif p in {"custom", "range"}:
        s, e = _to_date(date_from) or (today - timedelta(days=30)), _to_date(date_to) or today
    else:
        s, e = today.replace(day=1), today
    d = max(1, (e - s).days + 1)
    pe, ps = s - timedelta(days=1), s - timedelta(days=d)
    return s, e, ps, pe


def _pct(cur: int, prev: int) -> float:
    if prev == 0:
        return 100.0 if cur > 0 else 0.0
    return round(((cur - prev) / prev) * 100.0, 1)


def _serialize(obj: Any, cols: List[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for c in cols:
        v = getattr(obj, c, None)
        out[c] = v.isoformat() if isinstance(v, (date, datetime)) else v
    return out


def _search_filter(model: Any, needle: str):
    q = f"%{needle.lower()}%"
    clauses = []
    for c in ["client_name", "candidate_name", "recruiter_name", "am_name", "skill", "skill_set", "cp_name", "status", "invoice_no"]:
        if hasattr(model, c):
            clauses.append(func.lower(getattr(model, c)).like(q))
    return or_(*clauses) if clauses else True


def _distinct_non_empty(db: Session, model: Any, column_name: str) -> List[str]:
    col = getattr(model, column_name)
    rows = db.query(col).filter(col.isnot(None)).all()
    seen: Dict[str, str] = {}
    for (raw,) in rows:
        value = str(raw or "").strip()
        if not value:
            continue
        key = value.lower()
        if key not in seen:
            seen[key] = value
    return sorted(seen.values(), key=lambda x: x.lower())


def _norm_status(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _status_group_values(group: Optional[str]) -> Optional[set[str]]:
    if not group:
        return None
    key = _norm_status(group)
    return SUBMISSION_STATUS_GROUPS.get(key)


def _period_bounds(period: str, date_from: Optional[str], date_to: Optional[str]) -> Tuple[date, date]:
    start, end, _, _ = _window(period, date_from, date_to)
    return start, end


def _is_paid_status(value: Any) -> bool:
    return _norm_status(value) in {"paid", "settled", "completed"}


def _has_tracker_submissions(db: Session) -> bool:
    return db.query(models.TrackerSubmission.id).first() is not None


def _portal_submission_rows(
    db: Session,
    start: date,
    end: date,
    client: Optional[str] = None,
    am: Optional[str] = None,
    recruiter: Optional[str] = None,
):
    am_user = aliased(models.User)
    recruiter_user = aliased(models.User)
    q = (
        db.query(
            models.JobApplication,
            models.Job,
            am_user.full_name.label("am_full_name"),
            am_user.username.label("am_username"),
            am_user.email.label("am_email"),
            recruiter_user.full_name.label("recruiter_full_name"),
            recruiter_user.username.label("recruiter_username"),
            recruiter_user.email.label("recruiter_email"),
        )
        .join(models.Job, models.Job.id == models.JobApplication.job_id)
        .outerjoin(am_user, am_user.id == models.Job.account_manager_id)
        .outerjoin(recruiter_user, recruiter_user.id == models.JobApplication.recruiter_id)
        .filter(models.JobApplication.applied_at >= datetime.combine(start, datetime.min.time()))
        .filter(models.JobApplication.applied_at <= datetime.combine(end, datetime.max.time()))
    )
    if client:
        q = q.filter(func.lower(func.coalesce(models.Job.client_name, "")) == client.strip().lower())
    if am:
        needle = am.strip().lower()
        q = q.filter(
            func.lower(
                func.coalesce(
                    am_user.full_name,
                    am_user.username,
                    am_user.email,
                    "",
                )
            ) == needle
        )
    if recruiter:
        needle = recruiter.strip().lower()
        q = q.filter(
            func.lower(
                func.coalesce(
                    recruiter_user.full_name,
                    recruiter_user.username,
                    recruiter_user.email,
                    "",
                )
            ) == needle
        )
    raw_rows = q.all()

    # Fallback AM resolution from client ownership when jobs.account_manager_id is null.
    client_am_map: Dict[str, Tuple[str, str, str]] = {}
    client_names = {
        str(job.client_name or "").strip().lower()
        for _, job, *_ in raw_rows
        if str(job.client_name or "").strip()
    }
    if client_names:
        client_rows = (
            db.query(models.Client.client_name, models.User.full_name, models.User.username, models.User.email)
            .outerjoin(models.User, models.User.id == models.Client.am_id)
            .filter(func.lower(models.Client.client_name).in_(list(client_names)))
            .all()
        )
        for cname, full_name, username, email in client_rows:
            key = str(cname or "").strip().lower()
            if not key:
                continue
            client_am_map[key] = (
                str(full_name or "").strip(),
                str(username or "").strip(),
                str(email or "").strip(),
            )

    only_one_am: Tuple[str, str, str] = ("", "", "")
    am_users = (
        db.query(models.User.full_name, models.User.username, models.User.email)
        .filter(func.lower(func.coalesce(models.User.role, "")).in_(["account_manager", "account manager", "am"]))
        .all()
    )
    if len(am_users) == 1:
        full_name, username, email = am_users[0]
        only_one_am = (
            str(full_name or "").strip(),
            str(username or "").strip(),
            str(email or "").strip(),
        )

    out = []
    for app, job, am_full, am_user, am_email, rec_full, rec_user, rec_email in raw_rows:
        candidate = getattr(app, "candidate", None)
        if not any([str(am_full or "").strip(), str(am_user or "").strip(), str(am_email or "").strip()]):
            ckey = str(getattr(job, "client_name", None) or "").strip().lower()
            if ckey in client_am_map:
                mapped = client_am_map[ckey]
                if any(mapped):
                    am_full, am_user, am_email = mapped
                elif any(only_one_am):
                    am_full, am_user, am_email = only_one_am
            elif any(only_one_am):
                am_full, am_user, am_email = only_one_am
        out.append((app, job, candidate, am_full, am_user, am_email, rec_full, rec_user, rec_email))
    return out


def _portal_counts(rows):
    statuses = [_norm_status(getattr(app, "status", "")) for app, *_ in rows]
    total = len(statuses)
    shortlisted = sum(1 for s in statuses if s in PORTAL_SHORTLISTED)
    interviews = sum(1 for s in statuses if s in PORTAL_INTERVIEWS)
    selected = sum(1 for s in statuses if s in PORTAL_SELECTED)
    joined = sum(1 for s in statuses if s in PORTAL_JOINED)
    rejected = sum(1 for s in statuses if "reject" in s)
    on_hold = sum(1 for s in statuses if s in PORTAL_ON_HOLD or "hold" in s)
    duplicates = sum(1 for s in statuses if "duplicate" in s)
    return {
        "total": total,
        "shortlisted": shortlisted,
        "interviews": interviews,
        "selected": selected,
        "joined": joined,
        "rejected": rejected,
        "on_hold": on_hold,
        "duplicates": duplicates,
    }


def _pick_experience(app: Any, candidate: Any) -> Optional[Any]:
    values = [
        getattr(app, "experience_years", None),
        getattr(candidate, "experience_years", None) if candidate is not None else None,
        getattr(candidate, "relevant_experience_years", None) if candidate is not None else None,
        getattr(candidate, "experience", None) if candidate is not None else None,
    ]
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return v
    for j in [getattr(app, "parsed_resume", None), getattr(candidate, "parsed_resume", None) if candidate is not None else None, getattr(candidate, "parsed_data_json", None) if candidate is not None else None]:
        if isinstance(j, dict):
            for k in ["experience_years", "total_experience", "experience"]:
                vv = j.get(k)
                if vv is None:
                    continue
                if isinstance(vv, str) and not vv.strip():
                    continue
                return vv
    return None


def _fmt_experience(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return f"{float(value):g} yrs"
    s = str(value).strip()
    if not s:
        return None
    return s


def _pick_notice_period(app: Any, candidate: Any) -> Optional[str]:
    values = [
        getattr(candidate, "notice_period", None) if candidate is not None else None,
        getattr(candidate, "notice_period_days", None) if candidate is not None else None,
    ]
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            return f"{int(v)} days"
        s = str(v).strip()
        if s:
            return s
    for j in [getattr(app, "parsed_resume", None), getattr(candidate, "parsed_resume", None) if candidate is not None else None, getattr(candidate, "parsed_data_json", None) if candidate is not None else None]:
        if isinstance(j, dict):
            for k in ["notice_period", "notice_period_days"]:
                vv = j.get(k)
                if vv is None:
                    continue
                if isinstance(vv, (int, float)):
                    return f"{int(vv)} days"
                s = str(vv).strip()
                if s:
                    return s
    return None


@router.get("/filter-options")
def tracker_filter_options(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    legacy_defaults = [
        "CookieMan",
        "ITC Infotech",
        "ITC Ltd",
        "TCL",
        "TCTSL",
    ]

    existing_client_names = {
        (str(row[0]).strip().lower())
        for row in db.query(models.Client.client_name).all()
        if row and row[0]
    }
    to_create = []
    for legacy_name in legacy_defaults:
        if legacy_name.lower() not in existing_client_names:
            to_create.append(models.Client(client_name=legacy_name))
    if to_create:
        db.add_all(to_create)
        db.commit()

    clients: Dict[str, str] = {}
    ams: Dict[str, str] = {}
    recruiters: Dict[str, str] = {}

    def add_values(target: Dict[str, str], values: List[str]) -> None:
        for value in values:
            k = value.lower()
            if k not in target:
                target[k] = value

    add_values(clients, _distinct_non_empty(db, models.TrackerSubmission, "client_name"))
    add_values(clients, _distinct_non_empty(db, models.TrackerSelection, "client_name"))
    add_values(clients, _distinct_non_empty(db, models.TrackerClientInvoice, "client_name"))
    # Keep client dropdown aligned with Account Manager side:
    # include canonical client profile names and legacy job-linked client names.
    add_values(clients, _distinct_non_empty(db, models.Client, "client_name"))
    add_values(clients, _distinct_non_empty(db, models.Job, "client_name"))

    add_values(ams, _distinct_non_empty(db, models.TrackerSubmission, "am_name"))
    add_values(ams, _distinct_non_empty(db, models.TrackerSelection, "am_name"))
    # Also include real AM users so dropdown works even when tracker tables are empty.
    am_users = (
        db.query(models.User.full_name, models.User.username, models.User.email)
        .filter(
            func.lower(func.coalesce(models.User.role, "")).in_(
                ["account_manager", "account manager", "am"]
            )
        )
        .all()
    )
    am_names: List[str] = []
    for full_name, username, email in am_users:
        display = str(full_name or "").strip() or str(username or "").strip() or str(email or "").strip()
        if display:
            am_names.append(display)
    add_values(ams, am_names)

    add_values(recruiters, _distinct_non_empty(db, models.TrackerSubmission, "recruiter_name"))
    add_values(recruiters, _distinct_non_empty(db, models.TrackerSelection, "recruiter_name"))

    return {
        "clients": sorted(clients.values(), key=lambda x: x.lower()),
        "ams": sorted(ams.values(), key=lambda x: x.lower()),
        "recruiters": sorted(recruiters.values(), key=lambda x: x.lower()),
    }


@router.get("/submission-kpis")
def submission_kpis(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, ps, pe = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        cur_rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        prev_rows = _portal_submission_rows(db, ps, pe, client=client, am=am, recruiter=recruiter)
        m = _portal_counts(cur_rows)
        p = _portal_counts(prev_rows)
        return {"period": {"start": s.isoformat(), "end": e.isoformat()}, "metrics": m, "change_pct": {k: _pct(m[k], p[k]) for k in m}}

    def base(a: date, b: date):
        q = db.query(models.TrackerSubmission).filter(
            models.TrackerSubmission.submission_date >= a,
            models.TrackerSubmission.submission_date <= b,
        )
        if client:
            q = q.filter(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
        if am:
            q = q.filter(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
        if recruiter:
            q = q.filter(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
        return q

    cur, prev = base(s, e), base(ps, pe)

    def c(q, st=None):
        return q.count() if st is None else q.filter(models.TrackerSubmission.status.in_(list(st))).count()

    m = {
        "total": c(cur),
        "shortlisted": c(cur, SHORTLISTED),
        "interviews": c(cur, INTERVIEWS),
        "selected": c(cur, SELECTED),
        "joined": c(cur, {"Joined"}),
        "rejected": c(cur, REJECTED),
        "on_hold": c(cur, {"Req on Hold"}),
        "duplicates": c(cur, {"Duplicate"}),
    }
    p = {
        "total": c(prev),
        "shortlisted": c(prev, SHORTLISTED),
        "interviews": c(prev, INTERVIEWS),
        "selected": c(prev, SELECTED),
        "joined": c(prev, {"Joined"}),
        "rejected": c(prev, REJECTED),
        "on_hold": c(prev, {"Req on Hold"}),
        "duplicates": c(prev, {"Duplicate"}),
    }
    return {"period": {"start": s.isoformat(), "end": e.isoformat()}, "metrics": m, "change_pct": {k: _pct(m[k], p[k]) for k in m}}


@router.get("/selection-kpis")
def selection_kpis(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        joined = sum(1 for app, *_ in rows if _norm_status(app.status) in PORTAL_JOINED)
        return {
            "joined": joined,
            "total_gp": 0.0,
            "avg_gp_pct": 0.0,
            "bgv_pending": 0,
            "monthly_billing": 0.0,
        }

    q = db.query(models.TrackerSelection).filter(models.TrackerSelection.date_of_joining >= s, models.TrackerSelection.date_of_joining <= e)
    if client:
        q = q.filter(func.lower(models.TrackerSelection.client_name) == client.strip().lower())
    if am:
        q = q.filter(func.lower(models.TrackerSelection.am_name) == am.strip().lower())
    if recruiter:
        q = q.filter(func.lower(models.TrackerSelection.recruiter_name) == recruiter.strip().lower())
    rows = q.all()
    gp = [float(r.gp_percent) for r in rows if r.gp_percent is not None]
    return {
        "joined": sum(1 for r in rows if (r.status or "").strip().lower() == "joined"),
        "total_gp": round(sum(float(r.gp_value or 0) for r in rows), 2),
        "avg_gp_pct": round(sum(gp) / len(gp), 2) if gp else 0.0,
        "bgv_pending": sum(1 for r in rows if (r.status or "").strip().lower() == "joined" and not r.bgv_final_dt),
        "monthly_billing": round(sum(float(r.billing_per_day or 0) * 30 for r in rows), 2),
    }


@router.get("/invoice-kpis")
def invoice_kpis(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        inv = db.query(models.Invoice).filter(
            models.Invoice.created_at >= datetime.combine(s, datetime.min.time()),
            models.Invoice.created_at <= datetime.combine(e, datetime.max.time()),
        )
        if client:
            inv = inv.filter(func.lower(func.coalesce(models.Invoice.client_name, "")) == client.strip().lower())
        rows = inv.all()
        total = sum(float(r.amount or 0) for r in rows)
        paid = sum(float(r.amount or 0) for r in rows if _norm_status(r.status) in {"paid", "settled"})
        return {
            "total_invoiced": round(total, 2),
            "paid": round(paid, 2),
            "outstanding": round(total - paid, 2),
            "gst_total": 0.0,
            "cp_payable_pending": 0.0,
        }

    inv = db.query(models.TrackerClientInvoice).filter(models.TrackerClientInvoice.invoice_date >= s, models.TrackerClientInvoice.invoice_date <= e)
    if client:
        inv = inv.filter(func.lower(models.TrackerClientInvoice.client_name) == client.strip().lower())
    rows = inv.all()
    total = sum(float(r.total_inv_value or 0) for r in rows)
    paid = sum(float(r.total_inv_value or 0) for r in rows if (r.status or "").strip().lower() == "paid" or r.payment_date)
    gst = sum(float(r.gst_amount or 0) for r in rows)
    cp = db.query(models.TrackerCPInvoice).filter(models.TrackerCPInvoice.cp_inv_date >= s, models.TrackerCPInvoice.cp_inv_date <= e).all()
    cpp = sum(float(r.cp_inv_value or 0) for r in cp if (r.cp_payment_status or "").strip().lower() != "paid")
    return {"total_invoiced": round(total, 2), "paid": round(paid, 2), "outstanding": round(total - paid, 2), "gst_total": round(gst, 2), "cp_payable_pending": round(cpp, 2)}


@router.get("/submissions-by-client")
def submissions_by_client(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        out: Dict[str, Dict[str, Any]] = {}
        for app, job, *_ in rows:
            cname = str(getattr(job, "client_name", None) or "Unknown").strip() or "Unknown"
            status = str(getattr(app, "status", None) or "Unknown").strip() or "Unknown"
            if cname not in out:
                out[cname] = {"client_name": cname, "count": 0, "status_breakdown": {}}
            out[cname]["count"] += 1
            out[cname]["status_breakdown"][status] = out[cname]["status_breakdown"].get(status, 0) + 1
        return {"items": sorted(out.values(), key=lambda x: x["count"], reverse=True)}

    q = (
        db.query(models.TrackerSubmission.client_name, models.TrackerSubmission.status, func.count(models.TrackerSubmission.id))
        .filter(models.TrackerSubmission.submission_date >= s, models.TrackerSubmission.submission_date <= e)
    )
    if client:
        q = q.filter(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
    if am:
        q = q.filter(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
    if recruiter:
        q = q.filter(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
    rows = q.group_by(models.TrackerSubmission.client_name, models.TrackerSubmission.status).all()
    out: Dict[str, Dict[str, Any]] = {}
    for client_name, status, count in rows:
        key = client_name or "Unknown"
        if key not in out:
            out[key] = {"client_name": key, "count": 0, "status_breakdown": {}}
        out[key]["count"] += int(count or 0)
        out[key]["status_breakdown"][status or "Unknown"] = int(count or 0)
    return {"items": sorted(out.values(), key=lambda x: x["count"], reverse=True)}


@router.get("/status-funnel")
def status_funnel(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        c = _portal_counts(rows)
        total = c["total"]
        items = [
            {"status_group": "Submitted", "count": c["total"], "pct": round((c["total"] / total) * 100, 2) if total else 0.0},
            {"status_group": "Shortlisted", "count": c["shortlisted"], "pct": round((c["shortlisted"] / total) * 100, 2) if total else 0.0},
            {"status_group": "Interview", "count": c["interviews"], "pct": round((c["interviews"] / total) * 100, 2) if total else 0.0},
            {"status_group": "Selected", "count": c["selected"], "pct": round((c["selected"] / total) * 100, 2) if total else 0.0},
            {"status_group": "Joined", "count": c["joined"], "pct": round((c["joined"] / total) * 100, 2) if total else 0.0},
            {"status_group": "Rejected", "count": c["rejected"], "pct": round((c["rejected"] / total) * 100, 2) if total else 0.0},
        ]
        return {"items": items, "conversion_pct": round((c["joined"] / total) * 100, 2) if total else 0.0}

    q = db.query(models.TrackerSubmission).filter(models.TrackerSubmission.submission_date >= s, models.TrackerSubmission.submission_date <= e)
    if client:
        q = q.filter(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
    if am:
        q = q.filter(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
    if recruiter:
        q = q.filter(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
    total = q.count()
    items = []
    for label, st in [("Submitted", None), ("Shortlisted", SHORTLISTED), ("Interview", INTERVIEWS), ("Selected", SELECTED), ("Joined", {"Joined"}), ("Rejected", REJECTED)]:
        count = q.count() if st is None else q.filter(models.TrackerSubmission.status.in_(list(st))).count()
        items.append({"status_group": label, "count": count, "pct": round((count / total) * 100, 2) if total else 0.0})
    joined = next((i["count"] for i in items if i["status_group"] == "Joined"), 0)
    return {"items": items, "conversion_pct": round((joined / total) * 100, 2) if total else 0.0}


@router.get("/recruiter-leaderboard")
def recruiter_leaderboard(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        agg: Dict[str, Dict[str, Any]] = {}
        for app, _, _, _, _, _, rec_full, rec_user, rec_email in rows:
            name = str(rec_full or rec_user or rec_email or "Unknown").strip() or "Unknown"
            st = _norm_status(app.status)
            row = agg.setdefault(name, {"recruiter_name": name, "submissions": 0, "shortlisted": 0, "joined": 0, "avg_gp_pct": 0.0})
            row["submissions"] += 1
            if st in PORTAL_SHORTLISTED:
                row["shortlisted"] += 1
            if st in PORTAL_JOINED:
                row["joined"] += 1
        items = []
        for r in agg.values():
            subn = int(r["submissions"] or 0)
            jn = int(r["joined"] or 0)
            r["conversion_pct"] = round((jn / subn) * 100, 2) if subn else 0.0
            items.append(r)
        return {"items": sorted(items, key=lambda x: x["joined"], reverse=True)}

    q = (
        db.query(
            models.TrackerSubmission.recruiter_name,
            func.count(models.TrackerSubmission.id).label("submissions"),
            func.sum(case((models.TrackerSubmission.status.in_(list(SHORTLISTED)), 1), else_=0)).label("shortlisted"),
            func.sum(case((models.TrackerSubmission.status == "Joined", 1), else_=0)).label("joined"),
        )
        .filter(models.TrackerSubmission.submission_date >= s, models.TrackerSubmission.submission_date <= e)
    )
    if client:
        q = q.filter(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
    if am:
        q = q.filter(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
    if recruiter:
        q = q.filter(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
    rows = q.group_by(models.TrackerSubmission.recruiter_name).all()
    gp_map = {n: float(v or 0) for n, v in db.query(models.TrackerSelection.recruiter_name, func.avg(models.TrackerSelection.gp_percent)).group_by(models.TrackerSelection.recruiter_name).all()}
    items = []
    for n, sub, short, j in rows:
        subn = int(sub or 0)
        join = int(j or 0)
        items.append({"recruiter_name": n or "Unknown", "submissions": subn, "shortlisted": int(short or 0), "joined": join, "conversion_pct": round((join / subn) * 100, 2) if subn else 0.0, "avg_gp_pct": round(gp_map.get(n, 0.0), 2)})
    return {"items": sorted(items, key=lambda x: x["joined"], reverse=True)}


@router.get("/skills-breakdown")
def skills_breakdown(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        skill_counts: Dict[str, int] = {}
        accepted_counts: Dict[str, int] = {}
        for app, _, candidate, *_ in rows:
            status_key = _norm_status(app.status)
            raw_skills = getattr(app, "skills", None)
            parsed: List[str] = []
            if isinstance(raw_skills, list):
                parsed = [str(v).strip() for v in raw_skills if str(v).strip()]
            elif isinstance(raw_skills, str):
                parsed = [s.strip() for s in re.split(r"[,/|]", raw_skills) if s.strip()]
            if not parsed and candidate is not None:
                cskills = getattr(candidate, "skills", None)
                if isinstance(cskills, list):
                    parsed = [str(v).strip() for v in cskills if str(v).strip()]
                elif isinstance(cskills, str):
                    parsed = [s.strip() for s in re.split(r"[,/|]", cskills) if s.strip()]
            if not parsed:
                parsed = ["Unknown"]
            for sk in parsed:
                skill_counts[sk] = skill_counts.get(sk, 0) + 1
                if status_key in PORTAL_SELECTED or status_key in PORTAL_JOINED:
                    accepted_counts[sk] = accepted_counts.get(sk, 0) + 1
        ranked = sorted(skill_counts.items(), key=lambda kv: kv[1], reverse=True)[:10]
        return {"items": [{"skill": sk, "count": c, "acceptance_rate": round((accepted_counts.get(sk, 0) / c) * 100, 2) if c else 0.0} for sk, c in ranked]}

    q = (
        db.query(models.TrackerSubmission.skill, func.count(models.TrackerSubmission.id), func.sum(case((models.TrackerSubmission.status.in_(["Selected", "Joined"]), 1), else_=0)))
        .filter(models.TrackerSubmission.submission_date >= s, models.TrackerSubmission.submission_date <= e)
    )
    if client:
        q = q.filter(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
    if am:
        q = q.filter(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
    if recruiter:
        q = q.filter(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
    rows = q.group_by(models.TrackerSubmission.skill).order_by(desc(func.count(models.TrackerSubmission.id))).limit(10).all()
    return {"items": [{"skill": skill or "Unknown", "count": int(c or 0), "acceptance_rate": round((int(a or 0) / int(c or 1)) * 100, 2)} for skill, c, a in rows]}


@router.get("/am-performance")
def am_performance(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    s, e, _, _ = _window(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, s, e, client=client, am=am, recruiter=recruiter)
        agg: Dict[str, Dict[str, Any]] = {}
        for app, _, _, am_full, am_user, am_email, *_ in rows:
            name = str(am_full or am_user or am_email or "Unassigned").strip() or "Unassigned"
            st = _norm_status(app.status)
            row = agg.setdefault(name, {"am_name": name, "submissions": 0, "shortlisted": 0, "selections": 0, "total_gp": 0.0})
            row["submissions"] += 1
            if st in PORTAL_SHORTLISTED:
                row["shortlisted"] += 1
            if st in PORTAL_SELECTED or st in PORTAL_JOINED:
                row["selections"] += 1
        return {"items": list(agg.values())}

    q = (
        db.query(models.TrackerSubmission.am_name, func.count(models.TrackerSubmission.id), func.sum(case((models.TrackerSubmission.status.in_(list(SHORTLISTED)), 1), else_=0)), func.sum(case((models.TrackerSubmission.status.in_(list(SELECTED)), 1), else_=0)))
        .filter(models.TrackerSubmission.submission_date >= s, models.TrackerSubmission.submission_date <= e)
    )
    if client:
        q = q.filter(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
    if am:
        q = q.filter(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
    if recruiter:
        q = q.filter(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
    rows = q.group_by(models.TrackerSubmission.am_name).all()
    gp_map = {n: float(v or 0) for n, v in db.query(models.TrackerSelection.am_name, func.sum(models.TrackerSelection.gp_value)).group_by(models.TrackerSelection.am_name).all()}
    return {"items": [{"am_name": n or "Unknown", "submissions": int(su or 0), "shortlisted": int(sh or 0), "selections": int(se or 0), "total_gp": round(gp_map.get(n, 0.0), 2)} for n, su, sh, se in rows]}


def _paged(model: Any, page: int, limit: int, search: Optional[str], sort_by: str, sort_dir: str, filters: List[Any], db: Session, cols: List[str]):
    q = db.query(model)
    if filters:
        q = q.filter(and_(*filters))
    if search:
        q = q.filter(_search_filter(model, search))
    s_col = getattr(model, sort_by, None) or getattr(model, "created_at")
    q = q.order_by((asc if sort_dir.lower() == "asc" else desc)(s_col))
    total = q.count()
    rows = q.offset((page - 1) * limit).limit(limit).all()
    pages = max(1, (total + limit - 1) // limit) if total else 1
    return {"items": [_serialize(r, cols) for r in rows], "page": page, "limit": limit, "total": total, "total_pages": pages}


@router.get("/submissions")
def submissions_table(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=500),
    search: Optional[str] = Query(None),
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    status_group: Optional[str] = Query(None),
    sort_by: str = Query("submission_date"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    status_values = _status_group_values(status_group)
    start, end = _period_bounds(period, date_from, date_to)
    if not _has_tracker_submissions(db):
        rows = _portal_submission_rows(db, start, end, client=client, am=am, recruiter=recruiter)
        items = []
        for app, job, candidate, am_full, am_user, am_email, rec_full, rec_user, rec_email in rows:
            st = str(app.status or "")
            if status and _norm_status(st) != _norm_status(status):
                continue
            if status_values is not None and st not in status_values:
                continue
            raw_skills = app.skills
            if (not raw_skills) and candidate is not None:
                raw_skills = getattr(candidate, "skills", None)
            skill_text = ", ".join(raw_skills) if isinstance(raw_skills, list) else str(raw_skills or "")
            candidate_name = (
                str(getattr(candidate, "full_name", None) or "").strip()
                or str(app.full_name or "").strip()
                or "Unknown"
            )
            recruiter_name = (
                str(rec_full or "").strip()
                or str(rec_user or "").strip()
                or str(rec_email or "").strip()
                or "Unknown"
            )
            am_name = (
                str(am_full or "").strip()
                or str(am_user or "").strip()
                or str(am_email or "").strip()
                or "Unassigned"
            )
            item = {
                "serial_no": 0,
                "submission_date": app.applied_at.date().isoformat() if app.applied_at else None,
                "client_name": getattr(job, "client_name", None),
                "requirement_no": getattr(job, "job_id", None) or getattr(job, "title", None),
                "am_name": am_name,
                "recruiter_name": recruiter_name,
                "candidate_name": candidate_name,
                "skill": skill_text,
                "total_experience": _fmt_experience(_pick_experience(app, candidate)) or "Not Provided",
                "current_location": getattr(candidate, "current_location", None),
                "notice_period": _pick_notice_period(app, candidate) or "Not Provided",
                "status": st,
                "remarks": app.client_feedback,
            }
            if search:
                blob = " ".join(str(v or "") for v in item.values()).lower()
                if search.strip().lower() not in blob:
                    continue
            items.append(item)
        reverse = sort_dir.lower() != "asc"
        items = sorted(items, key=lambda x: str(x.get(sort_by) or ""), reverse=reverse)
        for idx, row in enumerate(items, start=1):
            row["serial_no"] = idx
        total = len(items)
        start = (page - 1) * limit
        paged = items[start:start + limit]
        pages = max(1, (total + limit - 1) // limit) if total else 1
        return {"items": paged, "page": page, "limit": limit, "total": total, "total_pages": pages}

    f = []
    if client: f.append(func.lower(models.TrackerSubmission.client_name) == client.strip().lower())
    if am: f.append(func.lower(models.TrackerSubmission.am_name) == am.strip().lower())
    if recruiter: f.append(func.lower(models.TrackerSubmission.recruiter_name) == recruiter.strip().lower())
    if status: f.append(func.lower(models.TrackerSubmission.status) == status.strip().lower())
    if status_values is not None: f.append(models.TrackerSubmission.status.in_(list(status_values)))
    f.append(models.TrackerSubmission.submission_date >= start)
    f.append(models.TrackerSubmission.submission_date <= end)
    return _paged(models.TrackerSubmission, page, limit, search, sort_by, sort_dir, f, db, ["id","serial_no","submission_date","client_name","requirement_no","am_name","recruiter_name","candidate_name","skill","total_experience","current_location","notice_period","status","remarks"])


@router.get("/selections")
def selections_table(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=500),
    search: Optional[str] = Query(None),
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    am: Optional[str] = Query(None),
    recruiter: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    bgv_pending: Optional[bool] = Query(None),
    sort_by: str = Query("date_of_joining"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    start, end = _period_bounds(period, date_from, date_to)
    f = []
    if client: f.append(func.lower(models.TrackerSelection.client_name) == client.strip().lower())
    if am: f.append(func.lower(models.TrackerSelection.am_name) == am.strip().lower())
    if recruiter: f.append(func.lower(models.TrackerSelection.recruiter_name) == recruiter.strip().lower())
    if status: f.append(func.lower(models.TrackerSelection.status) == status.strip().lower())
    if bgv_pending:
        f.append(func.lower(func.coalesce(models.TrackerSelection.status, "")) == "joined")
        f.append(models.TrackerSelection.bgv_final_dt.is_(None))
    f.append(models.TrackerSelection.date_of_joining >= start)
    f.append(models.TrackerSelection.date_of_joining <= end)
    return _paged(models.TrackerSelection, page, limit, search, sort_by, sort_dir, f, db, ["id","client_name","am_name","recruiter_name","candidate_name","skill_set","date_of_joining","billing_per_day","ctc_per_month","gp_value","gp_percent","status","bgv_with","bgv_type","bgv_initiated_dt","bgv_interim_dt","bgv_final_dt","po_no"])


@router.get("/channel-partners")
def channel_partners_table(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=500),
    search: Optional[str] = Query(None),
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    sort_by: str = Query("date_of_joining"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    start, end = _period_bounds(period, date_from, date_to)
    f = [func.lower(models.TrackerChannelPartner.status) == status.strip().lower()] if status else []
    f.append(models.TrackerChannelPartner.date_of_joining >= start)
    f.append(models.TrackerChannelPartner.date_of_joining <= end)
    result = _paged(models.TrackerChannelPartner, page, limit, search, sort_by, sort_dir, f, db, ["id","cp_name","candidate_name","skill_set","date_of_joining","cp_billing","routing_fee","infy_billing","status","bgv_with","po_no"])
    for item in result["items"]:
        cp_billing = float(item.get("cp_billing") or 0)
        routing_fee = float(item.get("routing_fee") or 0)
        infy_billing = float(item.get("infy_billing") or 0)
        item["margin"] = round(infy_billing - cp_billing - routing_fee, 2)
    return result


@router.get("/client-invoices")
def client_invoices_table(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=500),
    search: Optional[str] = Query(None),
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    sort_by: str = Query("invoice_date"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    start, end = _period_bounds(period, date_from, date_to)
    f = []
    if client: f.append(func.lower(models.TrackerClientInvoice.client_name) == client.strip().lower())
    if status: f.append(func.lower(models.TrackerClientInvoice.status) == status.strip().lower())
    if month: f.append(func.lower(models.TrackerClientInvoice.service_month) == month.strip().lower())
    f.append(models.TrackerClientInvoice.invoice_date >= start)
    f.append(models.TrackerClientInvoice.invoice_date <= end)
    result = _paged(models.TrackerClientInvoice, page, limit, search, sort_by, sort_dir, f, db, ["id","client_name","candidate_name","service_month","po_no","invoice_no","invoice_date","invoice_value","gst_amount","total_inv_value","status","payment_date"])
    today = datetime.utcnow().date()
    for item in result["items"]:
        inv_date = _to_date(item.get("invoice_date"))
        status_raw = item.get("status")
        is_overdue = bool(inv_date and (inv_date + timedelta(days=30) < today) and not _is_paid_status(status_raw))
        item["status_display"] = "Overdue" if is_overdue else (status_raw or "Pending")
    return result


@router.get("/cp-invoices")
def cp_invoices_table(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=500),
    search: Optional[str] = Query(None),
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    cp_name: Optional[str] = Query(None),
    payment_status: Optional[str] = Query(None),
    gst_status: Optional[str] = Query(None),
    sort_by: str = Query("cp_inv_date"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _must_access(current_user)
    start, end = _period_bounds(period, date_from, date_to)
    f = []
    if cp_name: f.append(func.lower(models.TrackerCPInvoice.cp_name) == cp_name.strip().lower())
    if payment_status: f.append(func.lower(models.TrackerCPInvoice.cp_payment_status) == payment_status.strip().lower())
    if gst_status: f.append(func.lower(models.TrackerCPInvoice.gst_status) == gst_status.strip().lower())
    f.append(models.TrackerCPInvoice.cp_inv_date >= start)
    f.append(models.TrackerCPInvoice.cp_inv_date <= end)
    result = _paged(models.TrackerCPInvoice, page, limit, search, sort_by, sort_dir, f, db, ["id","cp_name","candidate_name","service_month","client_inv_no","client_inv_value","client_payment_dt","cp_inv_no","cp_inv_date","cp_inv_value","cp_payment_status","gst_status","remarks"])
    today = datetime.utcnow().date()
    for item in result["items"]:
        cp_inv_value = float(item.get("cp_inv_value") or 0)
        paid = _is_paid_status(item.get("cp_payment_status"))
        cp_inv_date = _to_date(item.get("cp_inv_date"))
        overdue = bool(cp_inv_date and cp_inv_date + timedelta(days=30) < today and not paid)
        item["payment_status_display"] = "Paid" if paid else ("Overdue" if overdue else "Pending")
        item["net_payable"] = 0.0 if paid else round(cp_inv_value, 2)
    return result


def _section_from_text(text: str) -> Optional[str]:
    t = " ".join(str(text or "").lower().split())
    for k, v in SECTION_ALIAS.items():
        if k in t:
            return v
    return None


def _scan_sections(ws) -> Dict[str, List[Dict[str, Any]]]:
    out: Dict[str, List[Dict[str, Any]]] = {k: [] for k in SECTION_MODEL.keys()}
    r = 1
    while r <= ws.max_row:
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        key = _section_from_text(" ".join(str(v or "") for v in row))
        if not key:
            r += 1
            continue
        header = r + 1
        headers = [str(ws.cell(header, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        rr, empty = header + 1, 0
        while rr <= ws.max_row:
            vals = [ws.cell(rr, c).value for c in range(1, ws.max_column + 1)]
            if _section_from_text(" ".join(str(v or "") for v in vals)):
                break
            if not any(str(v or "").strip() for v in vals):
                empty += 1
                if empty >= 2:
                    break
                rr += 1
                continue
            empty = 0
            row_map: Dict[str, Any] = {}
            for idx, h in enumerate(headers):
                if idx >= len(vals) or not str(h).strip() or vals[idx] is None:
                    continue
                norm = re.sub(r"[^a-z0-9]+", "_", h.lower()).strip("_")
                val = vals[idx]
                target = None
                # light mapping by header fragments
                if key == "submissions":
                    m = {"date": "submission_date", "client": "client_name", "req": "requirement_no", "am": "am_name", "recruiter": "recruiter_name", "candidate": "candidate_name", "skill": "skill", "status": "status", "remark": "remarks", "s_": "serial_no", "exp": "total_experience", "location": "current_location", "notice": "notice_period", "spoc": "spoc_name"}
                elif key == "selections":
                    m = {"client": "client_name", "am": "am_name", "recruiter": "recruiter_name", "candidate": "candidate_name", "skill": "skill_set", "doj": "date_of_joining", "billing": "billing_per_day", "ctc": "ctc_per_month", "gp_percent": "gp_percent", "gp": "gp_value", "status": "status", "po": "po_no", "bgv_with": "bgv_with", "bgv_pre_post": "bgv_type", "initi": "bgv_initiated_dt", "interim": "bgv_interim_dt", "final": "bgv_final_dt", "remark": "remarks"}
                elif key == "channel_partners":
                    m = {"cp_name": "cp_name", "candidate": "candidate_name", "skill": "skill_set", "doj": "date_of_joining", "cp_billing": "cp_billing", "routing": "routing_fee", "infy_billing": "infy_billing", "status": "status", "po": "po_no", "bgv_with": "bgv_with", "remark": "remarks"}
                elif key == "client_invoices":
                    m = {"client": "client_name", "candidate": "candidate_name", "service": "service_month", "po": "po_no", "inv_date": "invoice_date", "inv_value": "invoice_value", "inv": "invoice_no", "gst": "gst_amount", "tot": "total_inv_value", "status": "status", "payment": "payment_date"}
                else:
                    m = {"cp_name": "cp_name", "candidate": "candidate_name", "service": "service_month", "client_inv": "client_inv_no", "client_inv_dt": "client_inv_date", "client_inv_val": "client_inv_value", "client_inv_payment": "client_payment_dt", "cp_inv_date": "cp_inv_date", "cp_inv_val": "cp_inv_value", "cp_inv": "cp_inv_no", "payment_status": "cp_payment_status", "gst_status": "gst_status", "remark": "remarks"}
                for mk, mv in m.items():
                    if mk in norm:
                        target = mv
                        break
                if not target:
                    continue
                if target.endswith("_date") or target.startswith("bgv_") or target in {"submission_date", "date_of_joining", "payment_date", "client_payment_dt", "cp_inv_date"}:
                    row_map[target] = _to_date(val)
                elif target in {"billing_per_day", "ctc_per_month", "gp_value", "gp_percent", "cp_billing", "routing_fee", "infy_billing", "invoice_value", "gst_amount", "total_inv_value", "client_inv_value", "cp_inv_value"}:
                    row_map[target] = _to_float(val)
                elif target == "serial_no":
                    try:
                        row_map[target] = int(float(val))
                    except Exception:
                        row_map[target] = None
                else:
                    row_map[target] = str(val).strip()
            if row_map:
                out[key].append(row_map)
            rr += 1
        r = rr
    return out


@router.post("/import")
@router.post("/submissions/import")
async def import_tracker(file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    _must_access(current_user)
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Upload .xlsx/.xlsm")
    wb = load_workbook(filename=BytesIO(await file.read()), data_only=True)
    all_rows: Dict[str, List[Dict[str, Any]]] = {k: [] for k in SECTION_MODEL.keys()}
    for ws in wb.worksheets:
        parsed = _scan_sections(ws)
        for key in all_rows:
            all_rows[key].extend(parsed[key])
    inserted = {}
    for key, rows in all_rows.items():
        model = SECTION_MODEL[key]
        c = 0
        for row in rows:
            db.add(model(**row))
            c += 1
        inserted[key] = c
    db.commit()
    return {"message": "Import successful", "records": inserted, "total_inserted": sum(inserted.values())}


@router.get("/export")
def export_tracker(section: str = Query(...), db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    _must_access(current_user)
    aliases = {"submission": "submissions", "submissions": "submissions", "selection": "selections", "selections": "selections", "cp": "channel_partners", "channel_partners": "channel_partners", "client_invoices": "client_invoices", "invoices": "client_invoices", "cp_invoices": "cp_invoices"}
    key = aliases.get((section or "").strip().lower())
    if not key:
        raise HTTPException(400, "Invalid section")
    model = SECTION_MODEL[key]
    rows = db.query(model).order_by(desc(model.created_at)).all()
    cols = [c.name for c in model.__table__.columns]
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(cols)
    for r in rows:
        data = _serialize(r, cols)
        w.writerow([data.get(c) for c in cols])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={key}.csv"})
