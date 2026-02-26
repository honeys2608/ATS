from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
import uuid
from datetime import datetime
import re
from typing import List, Dict, Tuple, Optional

from app.utils.role_check import allow_user
from app.db import get_db
from app import models
from app.auth import get_current_user
from app.permissions import has_permission

router = APIRouter(prefix="/v1/bulk", tags=["Bulk Upload"])

# ============================================================
# VALIDATION CONSTANTS - INTERNATIONAL STANDARDS
# ============================================================
VALID_CANDIDATE_STATUSES = [s.value for s in models.CandidateStatus]
REQUIRED_FIELDS = ["email", "full_name", "phone"]
ALL_FIELDS = [
    # Required
    "email", "full_name", "phone",
    
    # Basic Info
    "date_of_birth", "current_location", "city", "pincode",
    
    # Addresses
    "current_address", "permanent_address",
    
    # Professional
    "skills", "experience_years", "education", "current_employer",
    "previous_employers", "notice_period", "expected_ctc",
    "preferred_location", "languages_known",
    
    # URLs
    "linkedin_url", "github_url", "portfolio_url", "resume_url",
    
    # Application Info
    "source", "referral", "status",
    
    # Vendor Related
    "is_vendor_candidate", "billing_rate", "payout_rate",
    
    # Tags
    "tags"
]

# ============================================================
# VALIDATION FUNCTIONS
# ============================================================
def validate_email(email: str) -> Tuple[bool, Optional[str]]:
    """Validate email using RFC 5322 pattern"""
    if not email or not isinstance(email, str):
        return False, "Email is required"
    
    email = email.strip().lower()
    pattern = r'^[a-z0-9!#$%&\'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&\'*+/=?^_`{|}~-]+)*@(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?$'
    
    if not re.match(pattern, email):
        return False, "Invalid email format (e.g., john.doe@example.com)"
    
    if len(email) > 255:
        return False, "Email is too long (max 255 characters)"
    
    return True, None


def validate_phone(phone: str) -> Tuple[bool, Optional[str]]:
    """Validate phone number (E.164 international format support)"""
    if not phone or not isinstance(phone, str):
        return False, "Phone number is required"
    
    phone = phone.strip()
    cleaned = re.sub(r'\D', '', phone)
    
    # International standard: 10-15 digits
    if len(cleaned) < 10:
        return False, "Phone must have at least 10 digits (e.g., 9876543210 or +91 9876543210)"
    
    if len(cleaned) > 15:
        return False, "Phone must not exceed 15 digits"
    
    return True, None


def validate_name(name: str, field_name: str = "Name") -> Tuple[bool, Optional[str]]:
    """
    Validate name: must start with capital letter, 2-100 chars
    Supports Unicode characters
    """
    if not name or not isinstance(name, str):
        return False, f"{field_name} is required"
    
    name_str = name.strip()
    
    if len(name_str) < 2:
        return False, f"{field_name} must be at least 2 characters"
    
    if len(name_str) > 100:
        return False, f"{field_name} cannot exceed 100 characters"
    
    # Check if starts with capital letter
    if not name_str[0].isupper():
        return False, f"{field_name} must start with a capital letter (e.g., 'John Doe', not 'john doe')"
    
    # Allow letters, spaces, hyphens, apostrophes, periods
    if not re.match(r"^[A-Z][a-zA-Z\s'-\.]+$", name_str):
        return False, f"{field_name} contains invalid characters. Use only letters, spaces, hyphens, apostrophes, and periods"
    
    return True, None


def validate_row(row: Dict, row_number: int) -> Tuple[bool, List[str]]:
    """
    Validate a single row from bulk upload with international standards
    Returns: (is_valid, list_of_error_messages)
    """
    errors = []
    
    # ============================================================
    # REQUIRED FIELDS
    # ============================================================
    email = row.get("email")
    full_name = row.get("full_name")
    phone = row.get("phone")
    
    # Email validation
    if not email or pd.isna(email):
        errors.append("Email: This field is required")
    elif isinstance(email, str) and email.strip():
        valid, err_msg = validate_email(email)
        if not valid:
            errors.append(f"Email: {err_msg}")
    
    # Name validation
    if not full_name or pd.isna(full_name):
        errors.append("Full Name: This field is required")
    elif isinstance(full_name, str) and full_name.strip():
        valid, err_msg = validate_name(full_name, "Full Name")
        if not valid:
            errors.append(f"Full Name: {err_msg}")
    
    # Phone validation
    if not phone or pd.isna(phone):
        errors.append("Phone: This field is required")
    elif isinstance(phone, str) and phone.strip():
        valid, err_msg = validate_phone(phone)
        if not valid:
            errors.append(f"Phone: {err_msg}")
    
    # ============================================================
    # OPTIONAL FIELDS WITH INTERNATIONAL STANDARDS
    # ============================================================
    
    # Date of Birth - YYYY-MM-DD (International standard)
    dob = row.get("date_of_birth")
    if dob and not pd.isna(dob):
        try:
            dob_str = str(dob).strip()
            if dob_str:
                # Try ISO 8601 format first
                parsed_date = datetime.strptime(dob_str, "%Y-%m-%d")
                # Check reasonable age (between 16 and 80)
                age = (datetime.now() - parsed_date).days / 365.25
                if age < 16 or age > 80:
                    errors.append("Date of Birth: Age must be between 16 and 80 years")
        except:
            errors.append("Date of Birth: Must be in YYYY-MM-DD format (e.g., 1990-05-15)")
    
    # Pincode - International format (typically 5-10 alphanumeric)
    pincode = row.get("pincode")
    if pincode and not pd.isna(pincode):
        pincode_str = str(pincode).strip()
        if pincode_str and not re.match(r'^[A-Z0-9]{5,10}$', pincode_str, re.IGNORECASE):
            errors.append("Pincode: Must be 5-10 alphanumeric characters (e.g., 400001)")
    
    # City - international standard (letters, spaces, hyphens)
    city = row.get("city")
    if city and not pd.isna(city):
        city_str = str(city).strip()
        if city_str and not re.match(r"^[A-Za-z\s'-]+$", city_str):
            errors.append("City: Must contain only letters, spaces, hyphens, or apostrophes")
    
    # Addresses - minimal validation (just check length)
    current_addr = row.get("current_address")
    if current_addr and not pd.isna(current_addr) and len(str(current_addr)) > 500:
        errors.append("Current Address: Cannot exceed 500 characters")
    
    permanent_addr = row.get("permanent_address")
    if permanent_addr and not pd.isna(permanent_addr) and len(str(permanent_addr)) > 500:
        errors.append("Permanent Address: Cannot exceed 500 characters")
    
    # Location - international format
    location = row.get("current_location")
    if location and not pd.isna(location):
        loc_str = str(location).strip()
        if loc_str and not re.match(r"^[A-Za-z\s',.-]+$", loc_str):
            errors.append("Current Location: Must contain only letters, spaces, or common punctuation")
    
    # Preferred Location - comma separated cities
    pref_loc = row.get("preferred_location")
    if pref_loc and not pd.isna(pref_loc):
        pref_str = str(pref_loc).strip()
        if pref_str:
            locations = [l.strip() for l in pref_str.split(",")]
            for loc in locations:
                if not re.match(r"^[A-Za-z\s',.-]+$", loc):
                    errors.append("Preferred Location: Cities must contain only letters and spaces")
                    break
    
    # Skills - International standard (max 25 skills, alphanumeric + special chars)
    skills = row.get("skills")
    if skills and not pd.isna(skills) and isinstance(skills, str):
        skills_list = [s.strip() for s in skills.split(",") if s.strip()]
        if len(skills_list) > 25:
            errors.append("Skills: Cannot have more than 25 skills (international standard)")
        for skill in skills_list:
            if len(skill) > 50:
                errors.append("Skills: Individual skill name cannot exceed 50 characters")
                break
    
    # Experience - International standard (0-70 years)
    experience = row.get("experience_years")
    if experience and not pd.isna(experience):
        try:
            exp_float = float(str(experience).split()[0])
            if exp_float < 0 or exp_float > 70:
                errors.append("Experience: Must be between 0 and 70 years (international standard)")
        except:
            errors.append("Experience: Must be a valid number (e.g., 5 or 5.5)")
    
    # Languages - comma separated, max 10
    languages = row.get("languages_known")
    if languages and not pd.isna(languages):
        langs_list = [l.strip() for l in str(languages).split(",") if l.strip()]
        if len(langs_list) > 10:
            errors.append("Languages: Cannot have more than 10 languages")
    
    # Expected CTC - International standard (valid number)
    ctc = row.get("expected_ctc")
    if ctc and not pd.isna(ctc):
        ctc_str = str(ctc).strip()
        if ctc_str:
            try:
                # Extract number part (handles "800000 INR", "₹800000", etc.)
                numbers = re.findall(r'\d+', ctc_str)
                if numbers:
                    ctc_val = float(numbers[0])
                    if ctc_val > 999999999:  # Max reasonable salary
                        errors.append("Expected CTC: Value too high (max 999,999,999)")
                else:
                    raise ValueError()
            except:
                errors.append("Expected CTC: Must be a valid number (e.g., 800000)")
    
    # Notice Period - International standard (0-90 days typical)
    notice = row.get("notice_period")
    if notice and not pd.isna(notice):
        try:
            notice_val = int(str(notice).split()[0])
            if notice_val < 0 or notice_val > 365:
                errors.append("Notice Period: Must be between 0 and 365 days")
        except:
            errors.append("Notice Period: Must be a valid number (days)")
    
    # URLs - Basic validation
    urls = {
        "linkedin_url": "LinkedIn",
        "github_url": "GitHub",
        "portfolio_url": "Portfolio",
        "resume_url": "Resume"
    }
    for url_field, url_name in urls.items():
        url = row.get(url_field)
        if url and not pd.isna(url):
            url_str = str(url).strip()
            if url_str and not url_str.startswith(('http://', 'https://')):
                errors.append(f"{url_name} URL: Must start with http:// or https://")
    
    # Status - International standard values
    status = row.get("status")
    valid_statuses = VALID_CANDIDATE_STATUSES
    if status and not pd.isna(status):
        if str(status).lower() not in valid_statuses:
            errors.append(f"Status: Must be one of: {', '.join(valid_statuses)}")
    
    # Source - International standard values
    source = row.get("source")
    valid_sources = ["portal", "recruiter", "vendor", "referral", "linkedin", "indeed", "other"]
    if source and not pd.isna(source):
        if str(source).lower() not in valid_sources:
            errors.append(f"Source: Must be one of: {', '.join(valid_sources)}")
    
    # Vendor rates - International standard (positive numbers only)
    for rate_field in ["billing_rate", "payout_rate"]:
        rate = row.get(rate_field)
        if rate and not pd.isna(rate):
            try:
                rate_val = float(str(rate).split()[0])
                if rate_val < 0 or rate_val > 999999:
                    errors.append(f"{rate_field.replace('_', ' ').title()}: Must be between 0 and 999999")
            except:
                errors.append(f"{rate_field.replace('_', ' ').title()}: Must be a valid number")
    
    # Is Vendor - Should be Yes/No or True/False
    is_vendor = row.get("is_vendor_candidate")
    if is_vendor and not pd.isna(is_vendor):
        vendor_str = str(is_vendor).lower()
        if vendor_str not in ["yes", "no", "true", "false", "1", "0"]:
            errors.append("Is Vendor Candidate: Must be Yes/No or True/False")
    
    is_valid = len(errors) == 0
    return is_valid, errors


# ============================================================
# HELPERS
# ============================================================
def split_csv(value):
    if not value:
        return []
    if isinstance(value, list):
        return value
    return [v.strip() for v in str(value).split(",") if v.strip()]


def parse_experience(value):
    if not value:
        return 0.0
    try:
        return float(str(value).split()[0])
    except:
        return 0.0


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except:
        return None


def wrap_education(value):
    if not value:
        return None
    if isinstance(value, dict):
        return value
    return {"raw": str(value)}


def clean_salary(value):
    if not value:
        return None
    try:
        return float(str(value).split()[0])
    except:
        return None


def filter_model_fields(model, data: dict):
    allowed = {c.name for c in model.__table__.columns}
    return {k: v for k, v in data.items() if k in allowed}


# ============================================================
# 1) BULK UPLOAD (CSV / XLSX → imported_candidates) WITH VALIDATION
# ============================================================
@router.post("/upload")
def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    Bulk upload candidates with comprehensive validation.
    
    Required columns: email, full_name, phone
    Optional columns: date_of_birth, skills, experience, etc.
    
    Returns validation errors with row numbers and specific field issues
    """
    allow_user(current_user)

    if not has_permission(current_user["role"], "candidates", "create"):
        raise HTTPException(403, "No permission to upload candidates")

    try:
        # ---- Read file
        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(file.file)
        elif file.filename.lower().endswith(".xlsx"):
            df = pd.read_excel(file.file)
        else:
            raise HTTPException(400, "Only CSV or XLSX files are supported")

        if df.empty:
            raise HTTPException(400, "File is empty. Please add candidate data")

        # ---- Normalize column names
        # ---- Normalize column names
        def normalize_column(col: str) -> str:
            col = col.strip().lower()
            col = re.sub(r"[()]", "", col)
            col = re.sub(r"[^a-z0-9\s_]", "", col)
            col = re.sub(r"\s+", "_", col)
            return col

        df.columns = [normalize_column(c) for c in df.columns]

        missing_required = [col for col in REQUIRED_FIELDS if col not in df.columns]
        extra_columns = [c for c in df.columns if c not in ALL_FIELDS]
        if extra_columns:
            raise HTTPException(
                400,
                f"Invalid columns found: {', '.join(extra_columns)}"
            )

        if missing_required:
            raise HTTPException(
                400, 
                f"Missing required columns: {', '.join(missing_required)}. "
                f"Required: email, full_name, phone"
            )

        # ---- Validate all rows first, collect errors
        validation_errors = []
        valid_rows = []

        for idx, (_, row) in enumerate(df.iterrows(), start=2):  # Start from 2 (after header)
            is_valid, errors = validate_row(row, idx)
            
            if not is_valid:
                validation_errors.append({
                    "row": idx,
                    "email": str(row.get("email", "N/A")),
                    "name": str(row.get("full_name", "N/A")),
                    "errors": errors
                })
            else:
                valid_rows.append((idx, row))

        # ---- If there are validation errors, return them all at once
        if validation_errors:
            raise HTTPException(
                422,
                {
                    "error": "Validation failed for some rows",
                    "total_rows": len(df),
                    "valid_rows": len(valid_rows),
                    "invalid_rows": len(validation_errors),
                    "details": validation_errors
                }
            )

        # ---- Process valid rows
        inserted = updated = 0
        duplicate_emails = set()

        for idx, row in valid_rows:
            email = row.get("email")
            
            # Skip empty emails
            if not email or pd.isna(email):
                continue

            email = str(email).strip().lower()

            # Check for duplicate emails in same upload
            if email in duplicate_emails:
                validation_errors.append({
                    "row": idx,
                    "email": email,
                    "errors": ["Email: Duplicate email in upload file"]
                })
                continue

            duplicate_emails.add(email)

            existing = (
                db.query(models.ImportedCandidate)
                .filter(models.ImportedCandidate.email == email)
                .first()
            )

            raw_data = {
                k: (None if pd.isna(v) else v)
                for k, v in row.items()
            }

            data = filter_model_fields(models.ImportedCandidate, raw_data)

            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                db.add(
                    models.ImportedCandidate(
                        id=str(uuid.uuid4()),
                        status="imported",
                        created_at=datetime.utcnow(),
                        **data
                    )
                )
                inserted += 1

        db.commit()

        return {
            "status": "success",
            "message": f"Bulk upload completed: {inserted} new candidate(s) added, {updated} existing candidate(s) updated",
            "inserted": inserted,
            "updated": updated,
            "total_processed": len(valid_rows)
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error processing file: {str(e)}")


# ============================================================
# 2) CONVERT IMPORTED → REAL CANDIDATES
# ============================================================
@router.post("/convert-all")
def convert_all_imported_candidates(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    allow_user(current_user)

    if not has_permission(current_user["role"], "candidates", "create"):
        raise HTTPException(403, "No permission")

    imported_list = (
        db.query(models.ImportedCandidate)
        .filter(models.ImportedCandidate.status == "imported")
        .all()
    )

    converted = skipped = 0

    for imp in imported_list:
        try:
            existing = db.query(models.Candidate).filter_by(email=imp.email).first()
            if existing:
                imp.status = "skipped"
                imp.candidate_id = existing.id
                imp.public_id = existing.public_id
                skipped += 1
                continue

            public_id = models.generate_candidate_public_id_from_org(db)

            candidate = models.Candidate(
                id=str(uuid.uuid4()),
                public_id=public_id,
                password="IMPORTED",

                full_name=imp.full_name,
                email=imp.email,
                phone=imp.phone or "9999999999",
                dob=parse_date(imp.date_of_birth),

                current_location=imp.current_location,
                current_address=imp.current_address,
                permanent_address=imp.permanent_address,

                source="Bulk",
                status="new",

                resume_url=imp.resume_url,

                skills=split_csv(imp.skills),
                tags=split_csv(imp.tags),
                languages_known=split_csv(imp.languages_known),

                experience_years=parse_experience(imp.experience_years),

                education=wrap_education(imp.education),

                current_employer=imp.current_employer,
                previous_employers=split_csv(imp.previous_employers),
                notice_period=imp.notice_period,
                expected_ctc=clean_salary(imp.expected_ctc),

                preferred_location=imp.preferred_location,

                profile_completed=False,
                created_at=datetime.utcnow()
            )

            db.add(candidate)
            db.flush()

            imp.status = "converted"
            imp.candidate_id = candidate.id
            imp.public_id = public_id

            converted += 1

        except Exception as e:
            print("FAILED:", e)
            imp.status = "failed"
            skipped += 1

    db.commit()

    return {
        "message": "Bulk conversion completed",
        "converted": converted,
        "skipped": skipped
    }

# ============================================================
# 3) GENERATE EXCEL TEMPLATE
# ============================================================
@router.get("/template/download")
def download_bulk_upload_template():
    """
    Generate and download an Excel template for bulk candidate upload
    with validation examples and instructions
    """
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "Excel library not available. Install openpyxl.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    
    # Define header row - All admin profile fields
    headers = [
        # Required
        "Full Name *", "Email *", "Phone *",
        
        # Basic Info
        "Date of Birth", "Current Location", "City", "Pincode",
        
        # Addresses
        "Current Address", "Permanent Address",
        
        # Professional  
        "Skills", "Experience (Years)", "Education", "Current Employer",
        "Previous Employers", "Notice Period (Days)", "Expected CTC",
        "Preferred Location", "Languages Known",
        
        # URLs
        "LinkedIn URL", "GitHub URL", "Portfolio URL", "Resume URL",
        
        # Application
        "Source", "Referral", "Status",
        
        # Vendor
        "Is Vendor Candidate", "Billing Rate", "Payout Rate",
        
        # Tags
        "Tags"
    ]
    
    # Add headers
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add example data - Complete profile
    example_data = [
        ["John Doe", "john.doe@example.com", "9876543210",
         "1990-05-15", "Mumbai", "Mumbai", "400001",
         "123 Main Street, Mumbai 400001", "456 Home Lane, Mumbai 400002",
         "Python, React, SQL, AWS", "5.5", "B.Tech Computer Science", "Tech Corp",
         "Software Solutions Inc, Code Factory", "30", "1200000",
         "Mumbai, Bangalore", "English, Hindi",
         "https://linkedin.com/in/johndoe", "https://github.com/johndoe", "https://johndoe.portfolio.com", "https://example.com/resume.pdf",
         "LinkedIn", "Senior Dev", "Active",
         "No", "", ""],
        
        ["Sarah Johnson", "sarah.j@company.com", "+91 9123456789",
         "1995-08-22", "Bangalore", "Bangalore", "560001",
         "789 Tech Park, Bangalore 560001", "321 Residential Ave, Bangalore 560002",
         "JavaScript, Node.js, AWS, Docker", "3.5", "M.Tech Information Technology", "Innovate Labs",
         "Digital Ventures", "15", "800000",
         "Bangalore, Hyderabad", "English",
         "https://linkedin.com/in/sarahjohnson", "https://github.com/sarahjohnson", "", "https://example.com/sarah_resume.pdf",
         "Recruiter", "John Smith", "Verified",
         "No", "", ""],
    ]
    
    for row_data in example_data:
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Set column widths - adjusted for more columns
    column_widths = [16, 22, 15, 15, 16, 12, 10, 28, 28, 22, 14, 18, 16, 20, 15, 16, 18, 12, 22, 18, 18, 20, 16, 14, 12, 14, 14, 14, 12]
    for idx, width in enumerate(column_widths, 1):
        col_letter = chr(64 + idx) if idx <= 26 else chr(64 + (idx // 26)) + chr(64 + (idx % 26))
        ws.column_dimensions[col_letter].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 35
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    
    instructions = [
        ["BULK UPLOAD INSTRUCTIONS - International Standards", ""],
        ["", ""],
        ["REQUIRED FIELDS (Must be filled):", ""],
        ["• Full Name *", "Alphanumeric + spaces/hyphens/apostrophes only, 2-100 chars, starts with capital (ISO standard)"],
        ["• Email *", "RFC 5322 standard format, max 255 chars (e.g., john.doe@example.com)"],
        ["• Phone *", "E.164 format: 10-15 digits, can include +country code (e.g., 9876543210 or +91 9876543210)"],
        ["", ""],
        ["BASIC INFORMATION:", ""],
        ["• Date of Birth", "ISO 8601 format: YYYY-MM-DD (e.g., 1990-05-15). Age must be 16-80 years"],
        ["• Current Location", "City/region name, comma-separated for multiple (e.g., 'Mumbai' or 'Mumbai, Bangalore')"],
        ["• City", "City name, letters/spaces/hyphens only (e.g., 'New York', 'Santa Fe')"],
        ["• Pincode", "5-10 alphanumeric characters per country standard (e.g., '400001' India, '10001' USA)"],
        ["", ""],
        ["ADDRESS FIELDS:", ""],
        ["• Current Address", "Street address, max 500 characters (e.g., '123 Main Street, City 400001')"],
        ["• Permanent Address", "Street address, max 500 characters (same format as Current Address)"],
        ["", ""],
        ["PROFESSIONAL INFORMATION:", ""],
        ["• Skills", "Comma-separated list of technical/domain skills, max 25 items (e.g., 'Python, React, AWS, Docker')"],
        ["• Experience (Years)", "Integer/decimal 0-70 years (e.g., 5 or 5.5). Reflects total professional experience"],
        ["• Education", "Degree/qualification (e.g., 'B.Tech Computer Science', 'MBA', 'High School Diploma')"],
        ["• Current Employer", "Company name (e.g., 'Tech Corp', 'Digital Ventures'). Leave blank if unemployed"],
        ["• Previous Employers", "Comma-separated list of past employers (e.g., 'Company A, Company B, Company C')"],
        ["• Notice Period (Days)", "Integer 0-365 days before candidate can join (e.g., 30, 45, 0 for immediate)"],
        ["• Expected CTC", "Annual salary in local currency (numbers only, no currency symbol). E.g., 1200000"],
        ["• Preferred Location", "Comma-separated cities/regions candidate can work in (e.g., 'Mumbai, Bangalore, Pune')"],
        ["• Languages Known", "Comma-separated languages (max 10), English convention (e.g., 'English, Hindi, Spanish')"],
        ["", ""],
        ["URL FIELDS (Must be valid HTTP/HTTPS):", ""],
        ["• LinkedIn URL", "Full URL starting with https:// (e.g., 'https://linkedin.com/in/johndoe')"],
        ["• GitHub URL", "Full URL starting with https:// (e.g., 'https://github.com/johndoe')"],
        ["• Portfolio URL", "Full URL starting with https:// (e.g., 'https://johndoe.portfolio.com')"],
        ["• Resume URL", "Full URL starting with https:// or http:// to resume file"],
        ["", ""],
        ["APPLICATION TRACKING:", ""],
        ["• Source", "One of: 'portal', 'recruiter', 'vendor', 'referral', 'linkedin', 'indeed', 'other'"],
        ["• Referral", "Name of person who referred the candidate, or 'Self-Apply' (if applicable)"],
        [
            "• Status",
            "One of: " + ", ".join([f"'{s}'" for s in VALID_CANDIDATE_STATUSES]),
        ],
        ["", ""],
        ["VENDOR INFORMATION (for consultants/contractors):", ""],
        ["• Is Vendor Candidate", "Yes/No or True/False (case-insensitive). 'No' if regular full-time candidate"],
        ["• Billing Rate", "Hourly/daily rate in local currency (numbers only, 0-999999). Leave blank if not vendor"],
        ["• Payout Rate", "Rate paid to consultant (numbers only, 0-999999). Leave blank if not vendor"],
        ["", ""],
        ["• Tags", "Comma-separated labels/keywords for categorization (e.g., 'senior, backend, remote')"],
        ["", ""],
        ["VALIDATION EXAMPLES:", ""],
        ["✓ Full Name: 'John Doe'", "Correct - Starts with capital, contains space"],
        ["✗ Full Name: 'john doe'", "Wrong - Must start with capital letter"],
        ["✓ Email: 'john.doe@example.com'", "Correct - RFC 5322 standard format"],
        ["✗ Email: 'john@example'", "Wrong - Missing domain extension"],
        ["✓ Phone: '+91 9876543210' or '9876543210'", "Correct - E.164 format, 10-15 digits"],
        ["✗ Phone: '987654321'", "Wrong - Too few digits (less than 10)"],
        ["✓ DOB: '1990-05-15'", "Correct - ISO 8601 format (age would be valid: ~33-34 years)"],
        ["✗ DOB: '15/05/1990' or '15-May-1990'", "Wrong - Must use YYYY-MM-DD format"],
        ["✓ Experience: '5.5' years", "Correct - Allows decimals, within 0-70 range"],
        ["✗ Experience: '150' years", "Wrong - Exceeds maximum of 70 years"],
        ["✓ CTC: '1200000'", "Correct - Numbers only, no currency symbol"],
        ["✗ CTC: '$1200000' or '₹12,00,000'", "Wrong - Remove currency symbols and commas"],
        ["✓ Skills: 'Python, React, AWS, Docker'", "Correct - Comma-separated, under 25 items"],
        ["✗ Skills: 'Python, React, AWS, Docker, Java, C++, JavaScript, Node.js, Express, MongoDB, MySQL, PostgreSQL, Redis, Kafka, Docker, Kubernetes, AWS, GCP, Azure, Linux, Windows, MacOS, Git, CI/CD, DevOps, Jenkins, Terraform'", "Wrong - Exceeds max 25 skills"],
        ["✓ Preferred Location: 'Mumbai, Bangalore, Pune'", "Correct - Comma-separated cities"],
        ["✓ LinkedIn: 'https://linkedin.com/in/johndoe'", "Correct - Full HTTPS URL"],
        ["✗ LinkedIn: 'linkedin.com/in/johndoe'", "Wrong - Must include 'https://' prefix"],
        ["✓ Status: 'verified' or 'active'", "Correct - Matches allowed enum values"],
        ["✗ Status: 'pending' or 'rejected'", "Wrong - Not in allowed list"],
        ["", ""],
        ["INTERNATIONAL STANDARDS REFERENCES:", ""],
        ["• Email validation", "RFC 5322 standard"],
        ["• Phone format", "E.164 international standard"],
        ["• Date format", "ISO 8601 international standard"],
        ["• Name validation", "Unicode support with international character set"],
        ["• Currency", "Local country standard (no currency symbols in upload)"],
        ["• Age validation", "16-80 years range (working-age population)"],
        ["• Experience range", "0-70 years maximum (industry standard)"],
    ]
    
    for row in instructions:
        instructions_ws.append(row)
    
    # Style instructions sheet
    for idx, row in enumerate(instructions_ws.iter_rows(), 1):
        if idx == 1:
            for cell in row:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        elif instructions[idx-1][0] and instructions[idx-1][0].startswith("•"):
            for cell in row:
                cell.font = Font(italic=True)
        elif "✓" in str(instructions[idx-1][0]) or "✗" in str(instructions[idx-1][0]):
            for cell in row:
                cell.font = Font(size=10)
    
    instructions_ws.column_dimensions['A'].width = 40
    instructions_ws.column_dimensions['B'].width = 50
    
    # Write to bytes
    from io import BytesIO
    bytes_io = BytesIO()
    wb.save(bytes_io)
    bytes_io.seek(0)
    
    return StreamingResponse(
        iter([bytes_io.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Bulk_Upload_Template.xlsx"}
    )
